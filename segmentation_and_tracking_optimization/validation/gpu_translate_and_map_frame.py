#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
GPU translation + mapping for ONE frame (ACDC <-> Imaris) using:
1) FFT phase-correlation on binned point histograms (coarse tx,ty)
2) GPU Sinkhorn OT to refine ONLY tx,ty  (robust, handles different counts)
   NOTE: since you already used a *_corrected.csv (scaled to um), we LOCK sx=sy=1.0.
3) Hard 1-1 matching (greedy from kNN) to output per-Imaris dx,dy residuals

Output:
- Excel with Summary + TranslationTrace + MatchesFrame + Unmatched lists
- JSON with locked translation parameters

Important fixes vs your previous version:
- Remove accidental double phase-correlation call
- Do NOT optimize sx/sy (pre-scaled ACDC input); avoid scale-collapse
- Matching uses the same transform model as refinement (sx=sy=1.0)
"""

import argparse
import json
import os
import re
import math
import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ---- Torch GPU ----
try:
    import torch
    HAVE_TORCH = True
except Exception:
    HAVE_TORCH = False


# -----------------------------
# Utils
# -----------------------------
def clean_colname(c: str) -> str:
    return re.sub(r"\s+", " ", str(c).strip())


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())


def find_col_by_candidates(df: pd.DataFrame, candidates):
    nm = {_norm(c): c for c in df.columns}
    for cand in candidates:
        key = _norm(cand)
        if key in nm:
            return nm[key]
    return None


def pretty_excel(path: str):
    wb = load_workbook(path)
    for sh in wb.sheetnames:
        ws = wb[sh]
        if ws.max_row < 1:
            continue
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.freeze_panes = "A2"
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            maxlen = 0
            for r in range(1, min(ws.max_row, 250) + 1):
                v = ws.cell(r, col).value
                if v is None:
                    continue
                maxlen = max(maxlen, len(str(v)))
            ws.column_dimensions[letter].width = min(max(10, maxlen + 2), 70)
    wb.save(path)


def robust_stats(x: np.ndarray):
    x = np.asarray(x, float)
    x = x[np.isfinite(x)]
    if x.size == 0:
        return {"median": np.nan, "mean": np.nan, "p90": np.nan, "p99": np.nan}
    return {
        "median": float(np.median(x)),
        "mean": float(np.mean(x)),
        "p90": float(np.percentile(x, 90)),
        "p99": float(np.percentile(x, 99)),
    }


# -----------------------------
# Load ACDC (single frame)
# -----------------------------
def read_acdc_frame(acdc_csv: str, channel: str, frame_i: int):
    df = pd.read_csv(acdc_csv, sep=None, engine="python")
    df.columns = [clean_colname(c) for c in df.columns]

    fcol = find_col_by_candidates(df, ["frame_i", "frame", "Frame"])
    if fcol is None:
        raise KeyError("ACDC missing frame_i.")

    idcol = find_col_by_candidates(df, ["Cell_ID", "Cell ID", "cell_id"])
    if idcol is None:
        raise KeyError("ACDC missing Cell_ID.")

    xcol = find_col_by_candidates(df, [f"{channel}_PositionX_um", "PositionX_um", "PosX_um"])
    ycol = find_col_by_candidates(df, [f"{channel}_PositionY_um", "PositionY_um", "PosY_um"])
    if xcol is None or ycol is None:
        raise KeyError(f"ACDC missing {channel}_PositionX_um / {channel}_PositionY_um.")

    area_col = find_col_by_candidates(df, [f"{channel}_Area_um2", "Area_um2", "Area"])
    # area is optional

    df[fcol] = pd.to_numeric(df[fcol], errors="coerce")
    df[xcol] = pd.to_numeric(df[xcol], errors="coerce")
    df[ycol] = pd.to_numeric(df[ycol], errors="coerce")
    if area_col is not None:
        df[area_col] = pd.to_numeric(df[area_col], errors="coerce")

    sub = df[df[fcol] == frame_i].dropna(subset=[xcol, ycol]).copy()
    sub[idcol] = sub[idcol].astype(str)

    xy = sub[[xcol, ycol]].to_numpy(float)
    ids = sub[idcol].to_numpy(str)
    areas = sub[area_col].to_numpy(float) if area_col is not None else None
    return ids, xy, areas


# -----------------------------
# Load Imaris positions (single time)
# -----------------------------
def detect_imaris_position_sheet(imaris_xls: str):
    xf = pd.ExcelFile(imaris_xls, engine="xlrd")
    best_name, best_df, best_score = None, None, -1

    for sh in xf.sheet_names:
        if "position" not in sh.lower():
            continue
        if "track" in sh.lower():
            continue

        try:
            df = pd.read_excel(imaris_xls, sheet_name=sh, engine="xlrd")
        except Exception:
            continue

        df.columns = [clean_colname(c) for c in df.columns]
        cols = [_norm(c) for c in df.columns]
        score = 0
        for w in ["id", "time", "positionx", "positiony"]:
            if any(w in c for c in cols):
                score += 2
        if df.shape[0] > 1000:
            score += 1

        if score > best_score:
            best_score = score
            best_name, best_df = sh, df

    if best_df is None or best_score < 6:
        raise ValueError("Could not auto-detect Imaris per-object Position sheet.")
    return best_name, best_df


def read_imaris_time(imaris_xls: str, time_value: int):
    sh, df = detect_imaris_position_sheet(imaris_xls)
    df.columns = [clean_colname(c) for c in df.columns]

    id_col = find_col_by_candidates(df, ["ID", "Id", "Object ID", "Track ID"])
    t_col = find_col_by_candidates(df, ["Time", "Time Index", "time", "t"])
    x_col = None
    y_col = None
    for c in df.columns:
        nc = _norm(c)
        if x_col is None and "positionx" in nc and "mean" not in nc:
            x_col = c
        if y_col is None and "positiony" in nc and "mean" not in nc:
            y_col = c

    if id_col is None or t_col is None or x_col is None or y_col is None:
        raise KeyError(f"Imaris sheet '{sh}' missing needed columns.")

    area_col = None
    for c in df.columns:
        if "area" in c.lower() and ("mean" not in c.lower()):
            area_col = c
            break

    df[t_col] = pd.to_numeric(df[t_col], errors="coerce")
    df[x_col] = pd.to_numeric(df[x_col], errors="coerce")
    df[y_col] = pd.to_numeric(df[y_col], errors="coerce")
    if area_col is not None:
        df[area_col] = pd.to_numeric(df[area_col], errors="coerce")

    sub = df[df[t_col] == time_value].dropna(subset=[x_col, y_col]).copy()
    sub[id_col] = sub[id_col].astype(str)

    xy = sub[[x_col, y_col]].to_numpy(float)
    ids = sub[id_col].to_numpy(str)
    areas = sub[area_col].to_numpy(float) if area_col is not None else None
    return sh, ids, xy, areas


# -----------------------------
# 1) Coarse translation via phase correlation (GPU FFT)
# -----------------------------
def build_histogram(xy: np.ndarray, origin: np.ndarray, bin_um: float, grid_hw: tuple[int, int]):
    H, W = grid_hw
    ij = np.floor((xy - origin[None, :]) / bin_um).astype(int)
    mask = (ij[:, 0] >= 0) & (ij[:, 0] < W) & (ij[:, 1] >= 0) & (ij[:, 1] < H)
    ij = ij[mask]
    hist = np.zeros((H, W), dtype=np.float32)
    if ij.shape[0] > 0:
        # y -> row, x -> col
        np.add.at(hist, (ij[:, 1], ij[:, 0]), 1.0)
    return hist


def phase_correlation_tx_ty(
    ac_xy: np.ndarray,
    im_xy: np.ndarray,
    bin_um: float = 2.0,
    pad_factor: float = 1.2,
    device="cuda",
):
    """
    Returns coarse (tx,ty) where ACDC + (tx,ty) ~ Imaris.
    """
    if not HAVE_TORCH:
        raise RuntimeError("PyTorch not available. Install torch with CUDA.")
    if device.startswith("cuda") and not torch.cuda.is_available():
        device = "cpu"

    all_xy = np.vstack([ac_xy, im_xy])
    mn = np.min(all_xy, axis=0)
    mx = np.max(all_xy, axis=0)
    span = mx - mn

    # pad bbox a bit
    mn = mn - 0.1 * span
    mx = mx + 0.1 * span
    span = mx - mn

    W = int(math.ceil(span[0] / bin_um))
    H = int(math.ceil(span[1] / bin_um))

    Wp = int(math.ceil(W * pad_factor))
    Hp = int(math.ceil(H * pad_factor))

    def next_pow2(x):
        p = 1
        while p < x:
            p *= 2
        return p

    Wp = next_pow2(max(8, Wp))
    Hp = next_pow2(max(8, Hp))

    origin = mn.copy()
    Ha = build_histogram(ac_xy, origin, bin_um, (Hp, Wp))
    Hb = build_histogram(im_xy, origin, bin_um, (Hp, Wp))

    A = torch.from_numpy(Ha).to(device)
    B = torch.from_numpy(Hb).to(device)

    FA = torch.fft.rfft2(A)
    FB = torch.fft.rfft2(B)
    R = FA * torch.conj(FB)
    R = R / (torch.abs(R) + 1e-8)
    corr = torch.fft.irfft2(R, s=A.shape)

    peak = torch.argmax(corr)
    py = int((peak // corr.shape[1]).item())
    px = int((peak % corr.shape[1]).item())

    if px > corr.shape[1] // 2:
        px = px - corr.shape[1]
    if py > corr.shape[0] // 2:
        py = py - corr.shape[0]

    tx = -px * bin_um
    ty = -py * bin_um
    return float(tx), float(ty), {"bin_um": bin_um, "grid": (Hp, Wp)}


# -----------------------------
# 2) Refine translation ONLY using Sinkhorn OT (GPU)
# -----------------------------
def sinkhorn_refine_translate_only(
    ac_xy: np.ndarray,
    im_xy: np.ndarray,
    init_params: dict,
    iters_outer=200,
    iters_sink=80,
    eps=20.0,
    lr=0.05,
    trim_q=0.70,
    device="cuda",
):
    """
    Fit: B ≈ A + t   (translation only)
    using Sinkhorn OT + trimmed loss.

    Because your ACDC input is already scaled to um (via *_corrected.csv),
    we lock sx=sy=1.0 to avoid scale-collapse.
    """
    if not HAVE_TORCH:
        raise RuntimeError("PyTorch not available.")
    if device.startswith("cuda") and not torch.cuda.is_available():
        device = "cpu"

    A = torch.tensor(ac_xy, dtype=torch.float32, device=device)
    B = torch.tensor(im_xy, dtype=torch.float32, device=device)

    tx = torch.tensor([init_params.get("tx", 0.0)], device=device, requires_grad=True)
    ty = torch.tensor([init_params.get("ty", 0.0)], device=device, requires_grad=True)

    nA, nB = A.shape[0], B.shape[0]
    a = torch.full((nA,), 1.0 / nA, device=device)
    b = torch.full((nB,), 1.0 / nB, device=device)

    opt = torch.optim.Adam([tx, ty], lr=lr)
    trace = []

    for step in range(iters_outer):
        opt.zero_grad()

        Ash = torch.stack([A[:, 0] + tx, A[:, 1] + ty], dim=1)
        C = torch.cdist(Ash, B, p=2.0) ** 2

        K = torch.exp(-C / eps) + 1e-12
        u = torch.ones_like(a)
        v = torch.ones_like(b)
        for _ in range(iters_sink):
            u = a / (K @ v + 1e-12)
            v = b / (K.t() @ u + 1e-12)

        P = torch.diag(u) @ K @ torch.diag(v)

        row_mass = P.sum(dim=1) + 1e-12
        d = (P * C).sum(dim=1) / row_mass  # expected sqerr per A

        q = int(max(10, int(trim_q * d.numel())))
        d_sorted, _ = torch.sort(d)
        loss = d_sorted[:q].mean()

        loss.backward()
        opt.step()

        trace.append(
            {
                "iter": step,
                "tx": float(tx.detach().cpu().item()),
                "ty": float(ty.detach().cpu().item()),
                "sx": 1.0,
                "sy": 1.0,
                "trimmed_mean_sqerr": float(loss.detach().cpu().item()),
            }
        )

        if step > 10 and abs(trace[-1]["trimmed_mean_sqerr"] - trace[-2]["trimmed_mean_sqerr"]) < 1e-6:
            break

    out = trace[-1]
    params = {"tx": out["tx"], "ty": out["ty"], "sx": 1.0, "sy": 1.0}
    return params, pd.DataFrame(trace)


# -----------------------------
# 3) Hard matching after translation (per Imaris cell)
# -----------------------------
def hard_match_per_imaris(
    ac_ids,
    ac_xy,
    im_ids,
    im_xy,
    tx_ty,
    sx: float = 1.0,
    sy: float = 1.0,
    k_candidates: int = 10,
    max_dist_um: float = 50.0,
    device: str = "cuda",
):
    """
    For each Imaris point, find a unique ACDC point (1-1).
    Uses GPU cdist + topk candidate search, then CPU greedy to enforce uniqueness.

    Transform applied to ACDC before matching:
        A' = [sx * x + tx,  sy * y + ty]
    Residuals reported are:
        dx_resid = Imaris_x - A'_x
        dy_resid = Imaris_y - A'_y
    """
    if not HAVE_TORCH:
        raise RuntimeError("PyTorch not available.")
    if device.startswith("cuda") and not torch.cuda.is_available():
        device = "cpu"

    tx, ty = float(tx_ty[0]), float(tx_ty[1])
    sx = float(sx)
    sy = float(sy)

    A0 = torch.tensor(ac_xy, dtype=torch.float32, device=device)
    A = torch.stack([A0[:, 0] * sx + tx, A0[:, 1] * sy + ty], dim=1)
    B = torch.tensor(im_xy, dtype=torch.float32, device=device)

    nB = B.shape[0]
    if A.shape[0] == 0 or nB == 0:
        dfm = pd.DataFrame(
            columns=[
                "Imaris_index",
                "Imaris_ID",
                "Imaris_x_um",
                "Imaris_y_um",
                "ACDC_index",
                "ACDC_Cell_ID",
                "ACDC_x_raw_um",
                "ACDC_y_raw_um",
                "ACDC_x_shifted_um",
                "ACDC_y_shifted_um",
                "dx_resid_um",
                "dy_resid_um",
                "err_um",
                "err_um2",
                "pos_dist_um",
            ]
        )
        return (
            dfm,
            pd.DataFrame({"Unmatched_Imaris_ID": list(im_ids)}),
            pd.DataFrame({"Unmatched_ACDC_Cell_ID": list(ac_ids)}),
        )

    k = min(int(k_candidates), int(A.shape[0]))
    if k <= 0:
        raise RuntimeError("k_candidates resulted in k<=0 (no A points).")

    cand_list = []
    chunk = 256
    for s in range(0, nB, chunk):
        e = min(s + chunk, nB)
        D = torch.cdist(B[s:e], A, p=2.0)  # (m, nA)
        vals, idx = torch.topk(D, k=k, largest=False, dim=1)
        cand_list.append((vals.detach().cpu().numpy(), idx.detach().cpu().numpy(), s, e))

    triples = []
    for vals, idx, s, e in cand_list:
        for i in range(e - s):
            im_i = s + i
            for kk in range(k):
                d = float(vals[i, kk])
                if d <= max_dist_um:
                    triples.append((im_i, int(idx[i, kk]), d))

    triples.sort(key=lambda x: x[2])
    used_A = set()
    used_B = set()
    matches = []
    for im_i, ac_i, d in triples:
        if im_i in used_B or ac_i in used_A:
            continue
        used_B.add(im_i)
        used_A.add(ac_i)
        matches.append((im_i, ac_i, d))

    rows = []
    for im_i, ac_i, d in matches:
        ax_raw, ay_raw = float(ac_xy[ac_i, 0]), float(ac_xy[ac_i, 1])
        ax = ax_raw * sx + tx
        ay = ay_raw * sy + ty

        ix, iy = float(im_xy[im_i, 0]), float(im_xy[im_i, 1])
        dx = ix - ax
        dy = iy - ay

        rows.append(
            {
                "Imaris_index": int(im_i),
                "Imaris_ID": im_ids[im_i],
                "Imaris_x_um": ix,
                "Imaris_y_um": iy,
                "ACDC_index": int(ac_i),
                "ACDC_Cell_ID": ac_ids[ac_i],
                "ACDC_x_raw_um": ax_raw,
                "ACDC_y_raw_um": ay_raw,
                "ACDC_x_shifted_um": ax,
                "ACDC_y_shifted_um": ay,
                "dx_resid_um": dx,
                "dy_resid_um": dy,
                "err_um": float(math.sqrt(dx * dx + dy * dy)),
                "err_um2": float(dx * dx + dy * dy),
                "pos_dist_um": float(d),
            }
        )

    dfm = (
        pd.DataFrame(rows).sort_values("err_um")
        if rows
        else pd.DataFrame(
            columns=[
                "Imaris_index",
                "Imaris_ID",
                "Imaris_x_um",
                "Imaris_y_um",
                "ACDC_index",
                "ACDC_Cell_ID",
                "ACDC_x_raw_um",
                "ACDC_y_raw_um",
                "ACDC_x_shifted_um",
                "ACDC_y_shifted_um",
                "dx_resid_um",
                "dy_resid_um",
                "err_um",
                "err_um2",
                "pos_dist_um",
            ]
        )
    )

    unmatched_im = [im_ids[i] for i in range(len(im_ids)) if i not in used_B]
    unmatched_ac = [ac_ids[i] for i in range(len(ac_ids)) if i not in used_A]

    return (
        dfm,
        pd.DataFrame({"Unmatched_Imaris_ID": unmatched_im}),
        pd.DataFrame({"Unmatched_ACDC_Cell_ID": unmatched_ac}),
    )


# -----------------------------
# Main
# -----------------------------
def main():
    ap = argparse.ArgumentParser(description="GPU translation + mapping for one frame (ACDC vs Imaris).")
    ap.add_argument("--acdc", required=True, help="ACDC CSV (use your *_corrected.csv)")
    ap.add_argument("--imaris", required=True, help="Imaris .xls")
    ap.add_argument("--out", required=True, help="Output Excel .xlsx")
    ap.add_argument("--channel", default="NIR", help="ACDC channel prefix")
    ap.add_argument("--frame", type=int, default=48, help="Frame number you refer to (e.g. 48)")
    ap.add_argument("--frame-is-1based", action="store_true", help="If set, frame=48 means ACDC frame_i=47")
    ap.add_argument("--imaris-time-offset", type=int, default=1, help="Imaris_Time = ACDC_frame_i + offset")
    ap.add_argument("--bin-um", type=float, default=2.0, help="Histogram bin size (um) for phase-corr")
    ap.add_argument("--eps", type=float, default=20.0, help="Sinkhorn epsilon (bigger=softer)")
    ap.add_argument("--outer-iters", type=int, default=200, help="Translation refine steps")
    ap.add_argument("--sink-iters", type=int, default=80, help="Sinkhorn iters per step")
    ap.add_argument("--lr", type=float, default=0.05, help="Adam LR for translation refine")
    ap.add_argument("--trim-q", type=float, default=0.70, help="Trim fraction for loss (ignore worst)")
    ap.add_argument("--k-candidates", type=int, default=10, help="kNN candidates per Imaris point")
    ap.add_argument("--max-dist-um", type=float, default=50.0, help="Max distance gate for matching")
    ap.add_argument("--target-mean-um2", type=float, default=1.0, help="Target err_um2 threshold (for frac_under)")
    ap.add_argument("--device", default="cuda", help="cuda or cpu")
    # kept only for CLI compatibility (ignored)
    ap.add_argument("--reg-scale", type=float, default=0.0, help="(ignored) scale is locked to 1.0")
    args = ap.parse_args()

    if not HAVE_TORCH:
        raise RuntimeError("Install torch first (GPU build).")

    frame_i = args.frame - 1 if args.frame_is_1based else args.frame
    im_time = frame_i + args.imaris_time_offset

    ac_ids, ac_xy, _ = read_acdc_frame(args.acdc, args.channel, frame_i)
    im_sheet, im_ids, im_xy, _ = read_imaris_time(args.imaris, im_time)

    if ac_xy.shape[0] < 10 or im_xy.shape[0] < 10:
        raise RuntimeError(f"Too few points: ACDC={ac_xy.shape[0]} Imaris={im_xy.shape[0]}")

    # 1) coarse
    tx0, ty0, pc_meta = phase_correlation_tx_ty(ac_xy, im_xy, bin_um=args.bin_um, device=args.device)

    # 2) refine translation only
    params_ref, trace_df = sinkhorn_refine_translate_only(
        ac_xy,
        im_xy,
        init_params={"tx": tx0, "ty": ty0},
        iters_outer=args.outer_iters,
        iters_sink=args.sink_iters,
        eps=args.eps,
        lr=args.lr,
        trim_q=args.trim_q,
        device=args.device,
    )

    t_ref = np.array([params_ref["tx"], params_ref["ty"]], dtype=float)
    sx = 1.0
    sy = 1.0

    # 3) hard matching (per Imaris) with consistent model
    matches_df, un_im_df, un_ac_df = hard_match_per_imaris(
        ac_ids,
        ac_xy,
        im_ids,
        im_xy,
        t_ref,
        sx=sx,
        sy=sy,
        k_candidates=args.k_candidates,
        max_dist_um=args.max_dist_um,
        device=args.device,
    )

    # metrics
    err2 = matches_df["err_um2"].to_numpy(float) if len(matches_df) else np.array([])
    err = matches_df["err_um"].to_numpy(float) if len(matches_df) else np.array([])
    stats_err2 = robust_stats(err2)
    stats_err = robust_stats(err)

    frac_under = float(np.mean(err2 <= args.target_mean_um2)) if err2.size else 0.0

    dx = matches_df["dx_resid_um"].to_numpy(float) if len(matches_df) else np.array([])
    dy = matches_df["dy_resid_um"].to_numpy(float) if len(matches_df) else np.array([])
    stats_dx = robust_stats(dx)
    stats_dy = robust_stats(dy)

    summary = pd.DataFrame(
        [
            {
                "acdc_frame_i": int(frame_i),
                "imaris_time": int(im_time),
                "imaris_position_sheet_used": im_sheet,
                "coarse_tx_um": float(tx0),
                "coarse_ty_um": float(ty0),
                "refined_tx_um": float(t_ref[0]),
                "refined_ty_um": float(t_ref[1]),
                "refined_sx": 1.0,
                "refined_sy": 1.0,
                "n_acdc_points": int(ac_xy.shape[0]),
                "n_imaris_points": int(im_xy.shape[0]),
                "matches_1to1": int(matches_df.shape[0]),
                "match_gate_max_dist_um": float(args.max_dist_um),
                "k_candidates": int(args.k_candidates),
                "err_um_median": stats_err["median"],
                "err_um_p90": stats_err["p90"],
                "err_um_p99": stats_err["p99"],
                "err_um2_mean": stats_err2["mean"],
                "err_um2_median": stats_err2["median"],
                "frac_err_um2_under_target": frac_under,
                "target_err_um2": float(args.target_mean_um2),
                "dx_resid_median": stats_dx["median"],
                "dy_resid_median": stats_dy["median"],
                "note": "Residuals are Imaris - (ACDC + refined_translation). Scale locked (sx=sy=1.0) because input is *_corrected.csv.",
            }
        ]
    )

    # save excel
    with pd.ExcelWriter(args.out, engine="openpyxl") as w:
        summary.to_excel(w, index=False, sheet_name="Summary")
        trace_df.to_excel(w, index=False, sheet_name="TranslationTrace")
        matches_df.to_excel(w, index=False, sheet_name="MatchesFrame")
        un_im_df.to_excel(w, index=False, sheet_name="Unmatched_Imaris")
        un_ac_df.to_excel(w, index=False, sheet_name="Unmatched_ACDC")

    pretty_excel(args.out)

    # save json “locked parameters”
    json_path = os.path.splitext(args.out)[0] + "_locked_translation.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "acdc_frame_i": int(frame_i),
                "imaris_time": int(im_time),
                "imaris_sheet": im_sheet,
                "translate_x_um": float(t_ref[0]),
                "translate_y_um": float(t_ref[1]),
                "scale_x": 1.0,
                "scale_y": 1.0,
                "coarse_translate_x_um": float(tx0),
                "coarse_translate_y_um": float(ty0),
                "bin_um": float(args.bin_um),
                "sinkhorn_eps": float(args.eps),
                "trim_q": float(args.trim_q),
                "device_used": args.device,
                "metrics": {
                    "err_um": stats_err,
                    "err_um2": stats_err2,
                    "dx_resid_um": stats_dx,
                    "dy_resid_um": stats_dy,
                    "frac_err_um2_under_target": frac_under,
                    "target_err_um2": float(args.target_mean_um2),
                },
            },
            f,
            indent=2,
        )

    print(f"Saved Excel: {args.out}")
    print(f"Saved JSON:  {json_path}")
    print(f"Coarse translation:  tx0={tx0:.6f} um, ty0={ty0:.6f} um")
    print(f"Refined translation: tx={t_ref[0]:.6f} um, ty={t_ref[1]:.6f} um (sx=sy=1.0 locked)")
    print(
        f"Matches 1-1: {matches_df.shape[0]} | "
        f"err_um median/p90/p99 = {stats_err['median']:.3f} / {stats_err['p90']:.3f} / {stats_err['p99']:.3f}"
    )
    print(f"err_um2 mean={stats_err2['mean']:.3f} | fraction <= target({args.target_mean_um2}) = {frac_under:.3f}")
    print(f"Residual medians: dx={stats_dx['median']:.3f} um, dy={stats_dy['median']:.3f} um")


if __name__ == "__main__":
    main()
