#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
acdc_to_imaris_structure_2pass.py

Convert Cell-ACDC CSV -> Imaris-like multi-sheet Excel workbook using a crosswalk CSV.

What’s new vs the 1-pass version:
  1) Supports "second-pass" derived measurements (tracks/kinematics/aggregates).
     If a parameter is missing in the crosswalk, the script can still fill it
     using auto-derived columns (e.g., Speed, Velocity X/Y, Track Length, etc.).
  2) Supports backtick-quoted ACDC column names in formulas:
        (`bbox-3` - `bbox-1`) * sqrt(cell_area_um2/cell_area_pxl)
     will be rewritten automatically into:
        (COL("bbox-3") - COL("bbox-1")) * sqrt(cell_area_um2/cell_area_pxl)

Inputs:
  --acdc      Cell-ACDC results CSV
  --crosswalk Crosswalk CSV with columns:
                imaris_label, cell_acdc_label, formula
              (optional: unit, category)
  --out       Output Excel (.xlsx)

Example:
  python acdc_to_imaris_structure_2pass.py --acdc field1.csv --crosswalk crosswalk.csv --out acdc_imaris_like.xlsx

Crosswalk logic:
  - If 'formula' is not empty: it is evaluated row-wise (one value per ACDC row)
  - Else if 'cell_acdc_label' exists: column is copied directly from ACDC
  - Else: if the script knows how to auto-derive it (2nd pass), it fills it
  - Else: the sheet stays empty (header only) and you’ll see it in CrosswalkReport

Formula language (safe):
  - arithmetic: + - * / **  and parentheses
  - reference identifier-safe columns directly: Area_um2, volume_um3, ...
  - reference ANY column using COL("exact column name")
  - backticks are auto-translated to COL():  `bbox-3`  -> COL("bbox-3")
  - functions: sqrt(x), log(x), exp(x), abs(x), clip(x,a,b), where(cond,a,b)

Second-pass auto-derived columns (created from tracks/positions if possible):
  - velocity_x_um_s, velocity_y_um_s, velocity_z_um_s, speed_um_s
  - displacement2_um2, acceleration_um_s2
  - track_length_total_um, track_displacement_x_um, track_displacement_y_um, track_displacement_z_um
  - track_displacement_length_um, track_duration_s
  - track_speed_mean_um_s, track_speed_max_um_s, track_speed_min_um_s
  - track_velocity_mean_um_s, track_velocity_max_um_s, track_velocity_min_um_s
  - track_straightness

Notes:
  - Track grouping uses the detected ID column (prefer 'Cell_ID').
  - Time is taken from frame_i/time_*; if it's frame_i, it’s converted to seconds
    using --frame-interval (default 1.0 s per frame).
"""

import argparse
import ast
import re
from typing import Dict, List, Tuple, Optional

import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ----------------------------
# Hardcoded Imaris sheet specs
# ----------------------------
SHEET_SPECS = """Overall\tVariable | Value | Unit | Time
Acceleration\tValue | Unit | Category | Time | Parent | ID
Area\tValue | Unit | Category | Time | Parent | ID
Center of Homogeneous Mass\tCenter of Homogeneous Mass X | Center of Homogeneous Mass Y | Center of Homogeneous Mass Z | Unit | Category | Collection | Time | Parent | ID
Center of Image Mass Ch=1\tCenter of Image Mass X | Center of Image Mass Y | Center of Image Mass Z | Unit | Category | Channel | Collection | Time | Parent | ID
Displacement^2\tValue | Unit | Category | Time | Parent | ID
Ellipsoid Axis A\tEllipsoid Axis A X | Ellipsoid Axis A Y | Ellipsoid Axis A Z | Unit | Category | Collection | Time | Parent | ID
Ellipsoid Axis B\tEllipsoid Axis B X | Ellipsoid Axis B Y | Ellipsoid Axis B Z | Unit | Category | Collection | Time | Parent | ID
Ellipsoid Axis C\tEllipsoid Axis C X | Ellipsoid Axis C Y | Ellipsoid Axis C Z | Unit | Category | Collection | Time | Parent | ID
Ellipsoid Axis Length A\tValue | Unit | Category | Time | Parent | ID
Ellipsoid Axis Length B\tValue | Unit | Category | Time | Parent | ID
Ellipsoid Axis Length C\tValue | Unit | Category | Time | Parent | ID
Ellipticity (oblate)\tValue | Unit | Category | Time | Parent | ID
Ellipticity (prolate)\tValue | Unit | Category | Time | Parent | ID
Intensity Center Ch=1\tIntensity Center Ch=1 | Unit | Category | Channel | Collection | Time | Parent | ID
Intensity Max Ch=1\tIntensity Max Ch=1 | Unit | Category | Channel | Collection | Time | Parent | ID
Intensity Mean Ch=1\tIntensity Mean Ch=1 | Unit | Category | Channel | Collection | Time | Parent | ID
Intensity Median Ch=1\tIntensity Median Ch=1 | Unit | Category | Channel | Collection | Time | Parent | ID
Intensity Min Ch=1\tIntensity Min Ch=1 | Unit | Category | Channel | Collection | Time | Parent | ID
Intensity StdDev Ch=1\tIntensity StdDev Ch=1 | Unit | Category | Channel | Collection | Time | Parent | ID
Intensity Sum Ch=1\tIntensity Sum Ch=1 | Unit | Category | Channel | Collection | Time | Parent | ID
Intensity Variance Ch=1\tIntensity Variance Ch=1 | Unit | Category | Channel | Collection | Time | Parent | ID
Location X\tValue | Unit | Category | Time | Parent | ID
Location Y\tValue | Unit | Category | Time | Parent | ID
Location Z\tValue | Unit | Category | Time | Parent | ID
Position X\tValue | Unit | Category | Time | Parent | ID
Position Y\tValue | Unit | Category | Time | Parent | ID
Position Z\tValue | Unit | Category | Time | Parent | ID
Speed\tValue | Unit | Category | Time | Parent | ID
Time\tValue | Unit | Category | Time | Parent | ID
Track Displacement Length\tValue | Unit | Category | Time | Parent | ID
Track Duration\tValue | Unit | Category | Time | Parent | ID
Track ID\tValue | Unit | Category | Time | Parent | ID
Track Length\tValue | Unit | Category | Time | Parent | ID
Track Speed Mean\tValue | Unit | Category | Time | Parent | ID
Track Straightness\tValue | Unit | Category | Time | Parent | ID
Track Velocity Mean\tValue | Unit | Category | Time | Parent | ID
Velocity X\tValue | Unit | Category | Time | Parent | ID
Velocity Y\tValue | Unit | Category | Time | Parent | ID
Velocity Z\tValue | Unit | Category | Time | Parent | ID
BoundingBoxAA Length A\tValue | Unit | Category | Time | Parent | ID
BoundingBoxAA Length B\tValue | Unit | Category | Time | Parent | ID
BoundingBoxAA Length C\tValue | Unit | Category | Time | Parent | ID
BoundingBoxOO Length A\tValue | Unit | Category | Time | Parent | ID
BoundingBoxOO Length B\tValue | Unit | Category | Time | Parent | ID
BoundingBoxOO Length C\tValue | Unit | Category | Time | Parent | ID
Ellipsoid Length A\tValue | Unit | Category | Time | Parent | ID
Ellipsoid Length B\tValue | Unit | Category | Time | Parent | ID
Ellipsoid Length C\tValue | Unit | Category | Time | Parent | ID
Number of Components\tValue | Unit | Category | Time | Parent | ID
Number of Disconnected Components\tValue | Unit | Category | Time | Parent | ID
Number of Disconnected Components per Time Point\tValue | Unit | Category | Time | Parent | ID
Number of Surfaces\tValue | Unit | Category | Time | Parent | ID
Number of Triangles\tValue | Unit | Category | Time | Parent | ID
Number of Tracks\tValue | Unit | Category | Time | Parent | ID
Number of Voxels\tValue | Unit | Category | Time | Parent | ID
Sphericity\tValue | Unit | Category | Time | Parent | ID
Surface Area\tValue | Unit | Category | Time | Parent | ID
Track Displacement X\tValue | Unit | Category | Time | Parent | ID
Track Displacement Y\tValue | Unit | Category | Time | Parent | ID
Track Displacement Z\tValue | Unit | Category | Time | Parent | ID
Track Speed Max\tValue | Unit | Category | Time | Parent | ID
Track Speed Min\tValue | Unit | Category | Time | Parent | ID
Track Velocity Max\tValue | Unit | Category | Time | Parent | ID
Track Velocity Min\tValue | Unit | Category | Time | Parent | ID
Volume\tValue | Unit | Category | Time | Parent | ID
""".strip("\n")


# ----------------------------
# Helpers
# ----------------------------
def clean_colname(c: str) -> str:
    return re.sub(r"\s+", " ", str(c).strip())


def parse_specs(spec_text: str) -> List[Tuple[str, List[str]]]:
    out = []
    for line in spec_text.splitlines():
        line = line.strip()
        if not line:
            continue
        if "\t" not in line:
            raise ValueError(f"Bad spec line (missing tab): {line}")
        sheet, headers = line.split("\t", 1)
        cols = [h.strip() for h in headers.split("|")]
        out.append((sheet.strip(), cols))
    return out


def safe_sheet_name(name: str, used: set) -> str:
    """Excel sheet name <= 31 chars, unique."""
    base = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(name)).strip()
    if not base:
        base = "Sheet"
    base = base[:31]
    if base not in used:
        used.add(base)
        return base
    for i in range(2, 10000):
        suf = f"_{i}"
        cand = (base[: (31 - len(suf))] + suf)
        if cand not in used:
            used.add(cand)
            return cand
    raise RuntimeError("Too many duplicate sheet names.")


def measurement_headers(headers: List[str]) -> List[str]:
    """Headers before the first 'Unit' are considered measurement columns."""
    idx_unit = None
    for i, h in enumerate(headers):
        if h.strip().lower() == "unit":
            idx_unit = i
            break
    if idx_unit is None:
        return headers[:1]
    return headers[:idx_unit]


def detect_id_col(df: pd.DataFrame) -> str:
    cols = list(df.columns)
    for c in cols:
        if c.lower() == "cell_id":
            return c
    for c in cols:
        lc = c.lower()
        if "cell" in lc and "id" in lc:
            return c
    for c in cols:
        if c.lower() in ["id", "label"]:
            return c
    raise RuntimeError("Could not detect Cell ID column in ACDC. Expected something like 'Cell_ID'.")


def detect_time_col(df: pd.DataFrame) -> str:
    cols = list(df.columns)
    for c in cols:
        if c.lower() == "frame_i":
            return c
    for c in cols:
        if c.lower() == "time_hours":
            return c
    for c in cols:
        if c.lower() == "time_minutes":
            return c
    for c in cols:
        lc = c.lower()
        if "frame" in lc or lc == "time":
            return c
    raise RuntimeError("Could not detect time column in ACDC (frame_i / time_hours / time_minutes).")


def _series_or_nan(df: pd.DataFrame, colname: str) -> pd.Series:
    if colname in df.columns:
        return pd.to_numeric(df[colname], errors="coerce")
    return pd.Series(np.nan, index=df.index, dtype=float)


def _detect_xy_cols(df: pd.DataFrame):
    """
    Best-effort position columns detection (um).
    Preference order:
      <CHANNEL>_PositionX_um / <CHANNEL>_PositionY_um
      PositionX / PositionY
      x_centroid / y_centroid
      centroid-0 / centroid-1
    """
    cols = list(df.columns)
    # channel-based
    cand_x = [c for c in cols if re.search(r"_positionx_?um$", c, re.I)]
    cand_y = [c for c in cols if re.search(r"_positiony_?um$", c, re.I)]
    if cand_x and cand_y:
        return cand_x[0], cand_y[0], None

    def pick(names):
        for n in names:
            for c in cols:
                if c.lower() == n.lower():
                    return c
        return None

    x = pick(["PositionX", "x_centroid", "centroid-0", "local_centroid-0"])
    y = pick(["PositionY", "y_centroid", "centroid-1", "local_centroid-1"])
    z = pick(["PositionZ", "z_centroid", "centroid-2", "local_centroid-2"])
    return x, y, z


# ----------------------------
# Safe formula evaluation
# ----------------------------
def _rewrite_backticks(formula: str) -> str:
    """
    Replace `col-name` with COL("col-name") so formulas can reference columns with '-'.
    """
    def repl(m):
        name = m.group(1)
        return f'COL("{name}")'
    return re.sub(r"`([^`]+)`", repl, formula)


def eval_rowwise_formula(formula: str, df: pd.DataFrame) -> pd.Series:
    """
    Evaluate a formula returning a Series aligned with df rows.
    Allowed:
      - arithmetic: + - * / ** ()
      - names: columns that are valid identifiers (e.g., Area_um2)
      - COL("exact column name") for any column name (including with -)
      - functions: sqrt, log, exp, abs, clip, where
      - backticks: `bbox-3` -> COL("bbox-3")
    """
    formula = str(formula).strip()
    if not formula:
        raise ValueError("Empty formula")
    if formula.startswith("="):
        formula = formula[1:].strip()

    formula = _rewrite_backticks(formula)

    env = {}
    for c in df.columns:
        if re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", c):
            env[c] = pd.to_numeric(df[c], errors="coerce")

    def COL(name: str):
        """Fetch a column by *exact* name or a forgiving normalized match."""
        name = str(name)
        if name in df.columns:
            return pd.to_numeric(df[name], errors="coerce")
        # Fuzzy match: ignore case and non-alphanumerics
        def _norm(s: str) -> str:
            return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())
        key = _norm(name)
        nm = {_norm(c): c for c in df.columns}
        if key in nm:
            c = nm[key]
            return pd.to_numeric(df[c], errors="coerce")
        raise KeyError(f'COL(): column "{name}" not found in ACDC.')

    allowed_funcs = {
        "COL": COL,
        "sqrt": lambda x: np.sqrt(x),
        "log": lambda x: np.log(x),
        "exp": lambda x: np.exp(x),
        "abs": lambda x: np.abs(x),
        "clip": lambda x, a, b: np.clip(x, a, b),
        "where": lambda cond, a, b: np.where(cond, a, b),
        "nan": np.nan,
    }

    tree = ast.parse(formula, mode="eval")

    def eval_node(node):
        if isinstance(node, ast.Expression):
            return eval_node(node.body)
        if isinstance(node, ast.Constant):
            return node.value
        if isinstance(node, ast.Name):
            if node.id in env:
                return env[node.id]
            if node.id in allowed_funcs:
                return allowed_funcs[node.id]
            raise ValueError(f"Unknown name '{node.id}'")
        if isinstance(node, ast.UnaryOp):
            v = eval_node(node.operand)
            if isinstance(node.op, ast.USub):
                return -v
            if isinstance(node.op, ast.UAdd):
                return +v
            raise ValueError("Unary operator not allowed")
        if isinstance(node, ast.BinOp):
            a = eval_node(node.left)
            b = eval_node(node.right)
            if isinstance(node.op, ast.Add):
                return a + b
            if isinstance(node.op, ast.Sub):
                return a - b
            if isinstance(node.op, ast.Mult):
                return a * b
            if isinstance(node.op, ast.Div):
                return a / b
            if isinstance(node.op, ast.Pow):
                return a ** b
            if isinstance(node.op, ast.Mod):
                return a % b
            raise ValueError("Binary operator not allowed")
        if isinstance(node, ast.Call):
            fn = eval_node(node.func)
            if fn not in allowed_funcs.values():
                raise ValueError("Function not allowed")
            args = [eval_node(a) for a in node.args]
            kwargs = {kw.arg: eval_node(kw.value) for kw in node.keywords}
            return fn(*args, **kwargs)
        raise ValueError(f"Disallowed expression node: {type(node).__name__}")

    out = eval_node(tree)
    if isinstance(out, (int, float, np.number)):
        return pd.Series([out] * len(df), index=df.index, dtype=float)
    if isinstance(out, np.ndarray):
        return pd.Series(out, index=df.index)
    if isinstance(out, pd.Series):
        return out.reindex(df.index)
    return pd.to_numeric(pd.Series(out, index=df.index), errors="coerce")


# ----------------------------
# Crosswalk load
# ----------------------------
def load_crosswalk(path: str) -> pd.DataFrame:
    cw = pd.read_csv(path)
    cw.columns = [clean_colname(c) for c in cw.columns]
    colmap = {c.lower(): c for c in cw.columns}
    need = {"imaris_label", "cell_acdc_label", "formula"}
    missing = need - set(colmap.keys())
    if missing:
        raise RuntimeError(f"Crosswalk must contain columns {sorted(list(need))}. Found: {list(cw.columns)}")

    cw = cw.rename(columns={
        colmap["imaris_label"]: "imaris_label",
        colmap["cell_acdc_label"]: "cell_acdc_label",
        colmap["formula"]: "formula",
    })

    if "unit" in colmap:
        cw = cw.rename(columns={colmap["unit"]: "unit"})
    if "category" in colmap:
        cw = cw.rename(columns={colmap["category"]: "category"})

    cw["imaris_label"] = cw["imaris_label"].astype(str).str.strip()
    cw["cell_acdc_label"] = cw["cell_acdc_label"].fillna("").astype(str).str.strip()
    cw["formula"] = cw["formula"].fillna("").astype(str).str.strip()
    return cw


def crosswalk_lookup(cw: pd.DataFrame) -> Dict[str, Dict[str, str]]:
    out = {}
    for _, r in cw.iterrows():
        key = str(r["imaris_label"]).strip()
        if not key:
            continue
        out[key] = {
            "cell_acdc_label": str(r.get("cell_acdc_label", "")).strip(),
            "formula": str(r.get("formula", "")).strip(),
            "unit": str(r.get("unit", "")).strip() if "unit" in cw.columns else "",
            "category": str(r.get("category", "")).strip() if "category" in cw.columns else "",
        }
    return out


# ----------------------------
# Workbook formatting
# ----------------------------
def format_workbook(path_xlsx: str):
    wb = load_workbook(path_xlsx)
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        if ws.max_row >= 2 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            for row in range(1, min(ws.max_row, 200) + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 45)
    wb.save(path_xlsx)


# ----------------------------
# Second-pass: add derived columns
# ----------------------------
def ensure_helper_columns(df: pd.DataFrame, id_col: str, time_col: str, frame_interval_s: float) -> pd.DataFrame:
    df = df.copy()

    # normalized ID + time
    df["_IMARIS_ID_"] = df[id_col].astype(str)
    df["_IMARIS_TIME_"] = pd.to_numeric(df[time_col], errors="coerce")

    # seconds timeline (used for kinematics)
    tl = time_col.lower()
    if tl == "time_hours" or "hours" in tl:
        df["_TIME_SEC_"] = df["_IMARIS_TIME_"] * 3600.0
    elif tl == "time_minutes" or "minutes" in tl:
        df["_TIME_SEC_"] = df["_IMARIS_TIME_"] * 60.0
    elif "frame" in tl:
        df["_TIME_SEC_"] = df["_IMARIS_TIME_"] * float(frame_interval_s)
    else:
        # assume already seconds-like (or at least monotonic)
        df["_TIME_SEC_"] = df["_IMARIS_TIME_"].astype(float)

    # position columns
    xcol, ycol, zcol = _detect_xy_cols(df)
    if xcol is not None:
        df["pos_x_um"] = pd.to_numeric(df[xcol], errors="coerce")
    else:
        df["pos_x_um"] = np.nan
    if ycol is not None:
        df["pos_y_um"] = pd.to_numeric(df[ycol], errors="coerce")
    else:
        df["pos_y_um"] = np.nan
    if zcol is not None:
        df["pos_z_um"] = pd.to_numeric(df[zcol], errors="coerce")
    else:
        df["pos_z_um"] = 0.0  # 2D default

    # Kinematics per track
    # Sort within track to compute diffs
    df["_SORT_KEY_"] = np.arange(len(df))
    df_sorted = df.sort_values(by=["_IMARIS_ID_", "_TIME_SEC_", "_SORT_KEY_"]).copy()

    g = df_sorted.groupby("_IMARIS_ID_", sort=False)

    dx = g["pos_x_um"].diff()
    dy = g["pos_y_um"].diff()
    dz = g["pos_z_um"].diff()
    dt = g["_TIME_SEC_"].diff()

    dt_safe = dt.replace(0, np.nan)

    vx = dx / dt_safe
    vy = dy / dt_safe
    vz = dz / dt_safe

    step_dist = np.sqrt(dx**2 + dy**2 + dz**2)
    speed = step_dist / dt_safe

    # acceleration from speed changes
    accel = speed.groupby(df_sorted["_IMARIS_ID_"]).diff() / dt_safe

    df_sorted["velocity_x_um_s"] = vx
    df_sorted["velocity_y_um_s"] = vy
    df_sorted["velocity_z_um_s"] = vz
    df_sorted["speed_um_s"] = speed
    df_sorted["displacement2_um2"] = step_dist ** 2
    df_sorted["acceleration_um_s2"] = accel

    # Track-level aggregates (repeat on each row for that track)
    first_x = g["pos_x_um"].transform("first")
    first_y = g["pos_y_um"].transform("first")
    first_z = g["pos_z_um"].transform("first")
    last_x = g["pos_x_um"].transform("last")
    last_y = g["pos_y_um"].transform("last")
    last_z = g["pos_z_um"].transform("last")
    first_t = g["_TIME_SEC_"].transform("first")
    last_t = g["_TIME_SEC_"].transform("last")

    disp_x = last_x - first_x
    disp_y = last_y - first_y
    disp_z = last_z - first_z
    disp_len = np.sqrt(disp_x**2 + disp_y**2 + disp_z**2)

    # total path length per track = sum of step distances
    track_len_total = g.apply(lambda gg: np.nansum(np.sqrt((gg["pos_x_um"].diff())**2 +
                                                         (gg["pos_y_um"].diff())**2 +
                                                         (gg["pos_z_um"].diff())**2).values))
    track_len_total = track_len_total.reindex(df_sorted["_IMARIS_ID_"]).values
    df_sorted["track_length_total_um"] = track_len_total

    df_sorted["track_displacement_x_um"] = disp_x
    df_sorted["track_displacement_y_um"] = disp_y
    df_sorted["track_displacement_z_um"] = disp_z
    df_sorted["track_displacement_length_um"] = disp_len

    dur = (last_t - first_t)
    df_sorted["track_duration_s"] = dur

    # mean speed/velocity
    df_sorted["track_speed_mean_um_s"] = df_sorted["track_length_total_um"] / dur.replace(0, np.nan)
    df_sorted["track_velocity_mean_um_s"] = disp_len / dur.replace(0, np.nan)

    # straightness
    df_sorted["track_straightness"] = disp_len / df_sorted["track_length_total_um"].replace(0, np.nan)

    # speed/velocity extremes using per-step speed
    # (for Imaris, "velocity" extremes are usually similar; we compute from speed)
    max_speed = g["speed_um_s"].transform("max")
    min_speed = g["speed_um_s"].transform("min")
    df_sorted["track_speed_max_um_s"] = max_speed
    df_sorted["track_speed_min_um_s"] = min_speed
    df_sorted["track_velocity_max_um_s"] = max_speed
    df_sorted["track_velocity_min_um_s"] = min_speed

    # Put back in original row order
    df_out = df_sorted.sort_values("_SORT_KEY_").drop(columns=["_SORT_KEY_"])
    return df_out


# ----------------------------
# Auto-derived default formulas (2nd pass)
# ----------------------------
AUTO_DERIVED = {
    "Acceleration": 'COL("acceleration_um_s2")',
    "Speed": 'COL("speed_um_s")',
    "Displacement^2": 'COL("displacement2_um2")',
    "Velocity X": 'COL("velocity_x_um_s")',
    "Velocity Y": 'COL("velocity_y_um_s")',
    "Velocity Z": 'COL("velocity_z_um_s")',
    "Track Displacement Length": 'COL("track_displacement_length_um")',
    "Track Displacement X": 'COL("track_displacement_x_um")',
    "Track Displacement Y": 'COL("track_displacement_y_um")',
    "Track Displacement Z": 'COL("track_displacement_z_um")',
    "Track Duration": 'COL("track_duration_s")',
    "Track Length": 'COL("track_length_total_um")',
    "Track Speed Mean": 'COL("track_speed_mean_um_s")',
    "Track Speed Max": 'COL("track_speed_max_um_s")',
    "Track Speed Min": 'COL("track_speed_min_um_s")',
    "Track Velocity Mean": 'COL("track_velocity_mean_um_s")',
    "Track Velocity Max": 'COL("track_velocity_max_um_s")',
    "Track Velocity Min": 'COL("track_velocity_min_um_s")',
    "Track Straightness": 'COL("track_straightness")',
}


# ----------------------------
# Main conversion
# ----------------------------
def main():
    ap = argparse.ArgumentParser(description="Convert Cell-ACDC CSV into Imaris-like sheet structure using a crosswalk CSV (+ 2nd-pass derivations).")
    ap.add_argument("--acdc", required=True, help="Cell-ACDC results CSV")
    ap.add_argument("--crosswalk", required=True, help="Crosswalk CSV (imaris_label, cell_acdc_label, formula)")
    ap.add_argument("--out", default="acdc_imaris_like.xlsx", help="Output Excel workbook (.xlsx)")
    ap.add_argument("--category_default", default="Surface", help="Default Category value if not provided")
    ap.add_argument("--frame-interval", type=float, default=1.0, help="Seconds per frame (used when time column is frame_i). Default: 1.0")
    args = ap.parse_args()

    # Load ACDC
    acdc = pd.read_csv(args.acdc)
    acdc.columns = [clean_colname(c) for c in acdc.columns]

    id_col = detect_id_col(acdc)
    time_col = detect_time_col(acdc)
    acdc = ensure_helper_columns(acdc, id_col=id_col, time_col=time_col, frame_interval_s=args.frame_interval)

    # Load crosswalk
    cw_df = load_crosswalk(args.crosswalk)
    cw = crosswalk_lookup(cw_df)

    specs = parse_specs(SHEET_SPECS)

    used_sheet_names = set()
    sheet_name_map = []
    report_rows = []
    auto_derived_keys = set()

    # --- Materialize helper columns for Overall aggregation (optional) ---
    # Some Overall metrics are sums over per-object metrics (e.g., total voxels).
    # If the ACDC CSV does not already contain those per-object columns, we try
    # to *materialize* them from the crosswalk formulas so Overall aggregation can work.

    def _norm_colname(s: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

    def _find_col_fuzzy(df: pd.DataFrame, name: str) -> Optional[str]:
        name = str(name)
        if name in df.columns:
            return name
        key = _norm_colname(name)
        nm = {_norm_colname(c): c for c in df.columns}
        return nm.get(key, None)

    def _materialize_from_crosswalk(imaris_key: str, target_col: str):
        if target_col in acdc.columns:
            return
        if imaris_key not in cw:
            return
        cell_lab = str(cw[imaris_key].get("cell_acdc_label", "")).strip()
        formula = str(cw[imaris_key].get("formula", "")).strip()
        if not (cell_lab or formula):
            return
        try:
            if formula:
                acdc[target_col] = eval_rowwise_formula(formula, acdc)
            else:
                src = _find_col_fuzzy(acdc, cell_lab)
                if not src:
                    raise KeyError(f"ACDC column not found: {cell_lab}")
                acdc[target_col] = pd.to_numeric(acdc[src], errors="coerce")
            report_rows.append({
                "imaris_label": f"[materialize]{imaris_key}",
                "status": f"OK -> created ACDC col '{target_col}' for Overall aggregation",
            })
        except Exception as e:
            report_rows.append({
                "imaris_label": f"[materialize]{imaris_key}",
                "status": f"ERROR -> could not create '{target_col}': {e}",
            })

    _materialize_from_crosswalk("Number of Disconnected Components", "num_disconnected_components")
    _materialize_from_crosswalk("Number of Voxels", "num_voxels")
    _materialize_from_crosswalk("Number of Surfaces", "num_surfaces")
    _materialize_from_crosswalk("Number of Triangles", "num_triangles")

# Aggregated sheets (per time point)
    AGG_TIME_SHEETS = {
        "Number of Tracks": lambda d: d.groupby("_IMARIS_TIME_")["_IMARIS_ID_"].nunique(),
        "Number of Surfaces": lambda d: d.groupby("_IMARIS_TIME_")["_IMARIS_ID_"].count(),
    }
    # Disconnected components per time if available
    if "num_disconnected_components" in acdc.columns:
        AGG_TIME_SHEETS["Number of Disconnected Components per Time Point"] = lambda d: d.groupby("_IMARIS_TIME_")["num_disconnected_components"].sum()

    with pd.ExcelWriter(args.out, engine="openpyxl") as writer:
        for sheet, headers in specs:
            excel_sheet_name = safe_sheet_name(sheet, used_sheet_names)
            sheet_name_map.append({"spec_sheet": sheet, "excel_sheet": excel_sheet_name})

            headers = [h.strip() for h in headers]
            meas_cols = measurement_headers(headers)

            # ----------------
            # Overall sheet
            # ----------------
            if sheet == "Overall":
                overall_labels = [k for k in cw.keys() if k.startswith("Overall::")]
                out_rows = []

                # Some helpful defaults if crosswalk rows exist but formula empty
                total_tracks = float(acdc["_IMARIS_ID_"].nunique())
                total_surfaces = float(len(acdc))
                per_time_surfaces = acdc.groupby("_IMARIS_TIME_")["_IMARIS_ID_"].count()

                for lab in sorted(overall_labels):
                    var = lab.split("::", 1)[1]
                    cell_lab = cw[lab]["cell_acdc_label"].strip()
                    formula = cw[lab]["formula"].strip()

                    # If no formula provided, use smart defaults for common Overall variables
                    if not formula and not cell_lab:
                        if var.lower() == "number of tracks":
                            formula = "TOTAL_COUNT_TRACKS()"
                        elif var.lower() == "total number of surfaces":
                            formula = "TOTAL_COUNT_SURFACES()"
                        elif var.lower() == "number of disconnected components per time point":
                            if "num_disconnected_components" in acdc.columns:
                                formula = "PER_TIME_SUM(num_disconnected_components)"
                            else:
                                formula = ""
                        elif var.lower() == "total number of disconnected components":
                            if "num_disconnected_components" in acdc.columns:
                                formula = "TOTAL_SUM(num_disconnected_components)"
                            else:
                                formula = ""
                        elif var.lower() == "total number of voxels":
                            # best effort: sum any likely voxel count column
                            for c in acdc.columns:
                                if c.lower() in ["num_voxels", "n_voxels", "voxel_count", "number_of_voxels"]:
                                    cell_lab = c
                                    formula = f"TOTAL_SUM(COL(\"{c}\"))"
                                    break

                    f = formula.strip()

                    if f.upper() == "TOTAL_COUNT_TRACKS()":
                        out_rows.append([var, total_tracks, "", ""])
                        report_rows.append({"imaris_label": lab, "status": "OK (auto Overall: total tracks)"})
                        continue
                    if f.upper() == "TOTAL_COUNT_SURFACES()":
                        out_rows.append([var, total_surfaces, "", ""])
                        report_rows.append({"imaris_label": lab, "status": "OK (auto Overall: total surfaces)"})
                        continue
                    if f.upper() == "PER_TIME_COUNT_SURFACES()":
                        for t, v in per_time_surfaces.items():
                            out_rows.append([var, float(v), "", float(t)])
                        report_rows.append({"imaris_label": lab, "status": "OK (auto Overall: per-time surfaces)"})
                        continue

                    # Supported aggregate formulas:
                    # PER_TIME_COUNT(), TOTAL_COUNT(), PER_TIME_SUM(col), TOTAL_SUM(col)
                    # plus allow "col" shorthand
                    if not f and cell_lab:
                        f = f"TOTAL_SUM(COL(\"{cell_lab}\"))" if cell_lab in acdc.columns else f"TOTAL_SUM({cell_lab})"

                    def get_col_series(name: str):
                        name = str(name).strip()
                        c = _find_col_fuzzy(acdc, name)
                        if c is None:
                            raise KeyError(name)
                        return pd.to_numeric(acdc[c], errors="coerce")

                    def parse_arg_inside(s: str) -> str:
                        m = re.search(r"\((.*)\)$", s.strip())
                        return m.group(1).strip() if m else ""

                    if f.upper() == "PER_TIME_COUNT()":
                        per_time_count = acdc.groupby("_IMARIS_TIME_")["_IMARIS_ID_"].count()
                        for t, v in per_time_count.items():
                            out_rows.append([var, float(v), "", float(t)])
                        report_rows.append({"imaris_label": lab, "status": "OK (PER_TIME_COUNT)"})

                    elif f.upper() == "TOTAL_COUNT()":
                        out_rows.append([var, float(len(acdc)), "", ""])
                        report_rows.append({"imaris_label": lab, "status": "OK (TOTAL_COUNT)"})

                    elif f.upper().startswith("PER_TIME_SUM(") and f.endswith(")"):
                        arg = parse_arg_inside(f)
                        try:
                            if arg.startswith("COL("):
                                name = arg[4:-1].strip().strip('"').strip("'")
                                s = get_col_series(name)
                            else:
                                s = get_col_series(arg)
                        except Exception as e:
                            report_rows.append({"imaris_label": lab, "status": f"EMPTY (Overall missing column: {e}) | formula={formula}"})
                            continue
                        per = pd.concat([acdc["_IMARIS_TIME_"], s], axis=1).groupby("_IMARIS_TIME_")[s.name].sum()
                        for t, v in per.items():
                            out_rows.append([var, float(v) if pd.notna(v) else np.nan, "", float(t)])
                        report_rows.append({"imaris_label": lab, "status": f"OK (PER_TIME_SUM {arg})"})

                    elif f.upper().startswith("TOTAL_SUM(") and f.endswith(")"):
                        arg = parse_arg_inside(f)
                        try:
                            if arg.startswith("COL("):
                                name = arg[4:-1].strip().strip('"').strip("'")
                                s = get_col_series(name)
                            else:
                                s = get_col_series(arg)
                        except Exception as e:
                            report_rows.append({"imaris_label": lab, "status": f"EMPTY (Overall missing column: {e}) | formula={formula}"})
                            continue
                        out_rows.append([var, float(np.nansum(s.values)), "", ""])
                        report_rows.append({"imaris_label": lab, "status": f"OK (TOTAL_SUM {arg})"})

                    else:
                        report_rows.append({"imaris_label": lab, "status": f"EMPTY (unsupported Overall formula: {formula})"})

                df_overall = pd.DataFrame(out_rows, columns=headers) if out_rows else pd.DataFrame(columns=headers)
                df_overall.to_excel(writer, sheet_name=excel_sheet_name, index=False)
                continue

            # ----------------
            # Aggregated per-time sheets (second pass)
            # ----------------
            if sheet in AGG_TIME_SHEETS and sheet not in cw:
                # If user provided mapping for these, we respect it; if not, we auto-fill.
                try:
                    ser = AGG_TIME_SHEETS[sheet](acdc)
                    rows = []
                    for t, v in ser.items():
                        rows.append([float(v), "", "Overall", float(t), "", ""])
                    out_df = pd.DataFrame(rows, columns=headers)
                    out_df.to_excel(writer, sheet_name=excel_sheet_name, index=False)
                    report_rows.append({"imaris_label": sheet, "status": "OK (auto-derived per-time aggregate)"})
                except Exception as e:
                    pd.DataFrame(columns=headers).to_excel(writer, sheet_name=excel_sheet_name, index=False)
                    report_rows.append({"imaris_label": sheet, "status": f"EMPTY (aggregate failed: {e})"})
                continue

            # ----------------
            # Normal sheets (per-object rows)
            # ----------------
            computed = {}
            any_filled = False

            for mc in meas_cols:
                # key mapping rules:
                # - If single measurement col and it is "Value" -> use sheet name
                # - If single measurement col equals sheet name -> use sheet name (fixes Intensity Max Ch=1)
                # - Else -> sheet::col
                if len(meas_cols) == 1:
                    if mc.strip().lower() == "value" or mc.strip().lower() == sheet.strip().lower():
                        key = sheet
                    else:
                        key = f"{sheet}::{mc}"
                else:
                    key = f"{sheet}::{mc}"

                # Fetch from crosswalk, else use auto-derived if known
                cell_lab = ""
                formula = ""
                if key in cw:
                    cell_lab = cw[key]["cell_acdc_label"].strip()
                    formula = cw[key]["formula"].strip()
                else:
                    # Backward-compatibility:
                    # Some older crosswalks used "Sheet::Sheet" for single-measurement sheets (Intensity*).
                    alt_key = f"{sheet}::{mc}"
                    if alt_key in cw:
                        cell_lab = cw[alt_key]["cell_acdc_label"].strip()
                        formula = cw[alt_key]["formula"].strip()
                        key = alt_key
                mapping_present = bool(cell_lab) or bool(formula)

                # Only auto-derive if the crosswalk mapping is missing/empty
                if (not mapping_present) and (key in AUTO_DERIVED):
                    formula = AUTO_DERIVED[key]
                    auto_derived_keys.add(key)
                    report_rows.append({"imaris_label": key, "status": f"OK (auto-derived) | formula={formula}"})
                elif (not mapping_present) and (sheet in AUTO_DERIVED) and (key not in cw):
                    # Some sheets are keyed by sheet name (e.g. Speed)
                    formula = AUTO_DERIVED[sheet]
                    auto_derived_keys.add(key)
                    report_rows.append({"imaris_label": key, "status": f"OK (auto-derived) | formula={formula}"})
                elif not mapping_present:
                    report_rows.append({"imaris_label": key, "status": "EMPTY (no mapping)"})
                    continue

                # treat "cala" as placeholder
                if formula.lower().startswith("cala") and not cell_lab:
                    report_rows.append({"imaris_label": key, "status": "EMPTY (formula=cala)"})
                    continue

                try:
                    if formula:
                        s = eval_rowwise_formula(formula, acdc)
                        computed[mc] = s
                        any_filled = True
                        if key in cw:
                            report_rows.append({"imaris_label": key, "status": "OK (formula)"})
                    elif cell_lab:
                        if cell_lab not in acdc.columns:
                            report_rows.append({"imaris_label": key, "status": f"ERROR: ACDC col not found: {cell_lab}"})
                            continue
                        computed[mc] = pd.to_numeric(acdc[cell_lab], errors="coerce")
                        any_filled = True
                        report_rows.append({"imaris_label": key, "status": f"OK (direct: {cell_lab})"})
                    else:
                        report_rows.append({"imaris_label": key, "status": "EMPTY (no mapping)"})
                except Exception as e:
                    report_rows.append({"imaris_label": key, "status": f"ERROR computing: {e} | formula={formula}"})

            if not any_filled:
                pd.DataFrame(columns=headers).to_excel(writer, sheet_name=excel_sheet_name, index=False)
                continue

            out_df = pd.DataFrame(index=acdc.index)
            for mc in meas_cols:
                out_df[mc] = computed.get(mc, np.nan)

            # metadata cols
            for h in headers[len(meas_cols):]:
                hh = h.strip()
                if hh.lower() == "unit":
                    out_df[hh] = ""
                elif hh.lower() == "category":
                    out_df[hh] = args.category_default
                elif hh.lower() == "channel":
                    out_df[hh] = "1" if "Ch=1" in sheet else ""
                elif hh.lower() == "collection":
                    out_df[hh] = ""
                elif hh.lower() == "time":
                    out_df[hh] = acdc["_IMARIS_TIME_"]
                elif hh.lower() == "id":
                    out_df[hh] = acdc["_IMARIS_ID_"]
                elif hh.lower() == "parent":
                    out_df[hh] = ""
                else:
                    out_df[hh] = ""

            out_df = out_df[headers]
            out_df = out_df.dropna(subset=["Time", "ID"], how="any")
            out_df.to_excel(writer, sheet_name=excel_sheet_name, index=False)

        pd.DataFrame(sheet_name_map).to_excel(writer, sheet_name="Index", index=False)
        pd.DataFrame(report_rows).to_excel(writer, sheet_name="CrosswalkReport", index=False)

        info = pd.DataFrame([
            ["acdc_file", args.acdc],
            ["crosswalk_file", args.crosswalk],
            ["detected_id_col", id_col],
            ["detected_time_col", time_col],
            ["frame_interval_s", args.frame_interval],
            ["note", "Sheets with no mapping stay empty. Some are auto-derived (2nd pass) if possible."],
        ], columns=["Key", "Value"])
        info.to_excel(writer, sheet_name="Info", index=False)

    format_workbook(args.out)
    print(f"[OK] Wrote Imaris-like workbook: {args.out}")


if __name__ == "__main__":
    main()
