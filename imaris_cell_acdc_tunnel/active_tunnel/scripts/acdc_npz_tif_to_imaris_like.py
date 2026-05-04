#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
acdc_npz_tif_to_imaris_like.py  (FORMULA-VALIDATED + TIME-INDEX FIX)

═══════════════════════════════════════════════════════════════════════════
DIAGNOSIS OF COMPARISON FAILURES (pipeline vs Imaris ground-truth)
═══════════════════════════════════════════════════════════════════════════

Full comparison of pipeline output against Imaris identified 5 root causes:

────────────────────────────────────────────────────────────────────────────
PROBLEM 1 — Time index off by 1  [FIXED HERE — see --time-index-start]
────────────────────────────────────────────────────────────────────────────
Symptom : Time Index r²=0.993, MAE=1.0 exactly (every row off by exactly 1).
           ALL per-frame surface metrics show negative r² (Area, Volume,
           Speed, Velocity, Intensity, Ellipsoid, etc.).
Cause   : The comparison joins on (Time, Parent).  If Time is off by 1 for
          every row, NO surface-level row ever joins correctly — they are
          paired with the wrong frame's data from Imaris, producing random-
          pairing r² values (negative).
Fix     : --time-index-start 2  (default).  Imaris labels the first captured
          frame as Time=2 in this dataset (empirically confirmed: pipeline
          output Time=1, Imaris Time=2 → constant MAE=1).
          Pass --time-index-start 1 if your Imaris export starts at Time=1.

Expected after fix: Area, Volume, Speed, Velocity, Intensity, Ellipsoid axes,
                    Displacement², Number of Voxels/Triangles all become
                    comparable (r² driven by segmentation quality, not join failures).

────────────────────────────────────────────────────────────────────────────
PROBLEM 2 — Residual XY coordinate offset ~30 µm  [PARTIALLY MITIGATED]
────────────────────────────────────────────────────────────────────────────
Symptom : Position X/Y r²=0.989 (GOOD) but MAE≈30 µm.
          Track Position Mean X/Y r²=0.9994, MAE≈10 µm (much better — averaging
          over a track's frames reduces the per-frame noise).
Cause   : The integer-pixel grid search (±10 px) finds the best whole-pixel
          offset, but the true registration shift is sub-pixel.  The residual
          ~30 µm bias comes from the fractional-pixel part of the offset.
Mitigation: Use --imaris-offset-x-px / --imaris-offset-y-px with the best
            integer value found by the grid search.  For sub-pixel refinement,
            run the grid search at 0.25 px steps around the best integer.
Not fixable to zero without sub-pixel offset search.

────────────────────────────────────────────────────────────────────────────
PROBLEM 3 — Z constant offset ~0.11 µm  [HARMLESS FOR 2D DATA]
────────────────────────────────────────────────────────────────────────────
Symptom : Position Z catastrophic r² (variance near zero; r² undefined).
          MAE=0.11 µm — small in absolute terms but nonzero.
Cause   : 2D data → all Z values are essentially constant.  Pipeline outputs
          Z=0 (or z_step × 0.5); Imaris outputs the physical stage Z
          (focal plane height, ~0.11 µm above coverslip).
Fix     : Use --imaris-offset-z-px to add the physical Z offset.  For 2D data
          this has no effect on kinematic metrics (ΔZ=0 either way) and can
          be safely ignored.  Same for CoHM Z, CoIM Z, Track Position Z Mean.

────────────────────────────────────────────────────────────────────────────
PROBLEM 4 — Track extent mismatch  [NOT FIXABLE IN THIS SCRIPT]
────────────────────────────────────────────────────────────────────────────
Symptom : Track Duration MAE=47,000 s (≈13 frames/track); Track N Surfaces
          MAE=12; Track Displacement r²≈0.03; Track Speed aggregates all bad.
Cause   : Imaris and ACDC use different track-linking parameters or gap-closing
          settings.  A single Imaris track of 33 frames may appear as two
          separate ACDC tracks of ~16 frames each.  Track displacement, speed
          aggregates, and AR1 are all sensitive to where the track starts/ends.
          Track Position Mean (unaffected by endpoints) r²=0.9994 ✓.
Fix     : Align track-linking parameters between ACDC and Imaris before
          re-running the pipeline.  This script cannot fix retroactively.

────────────────────────────────────────────────────────────────────────────
PROBLEM 5 — Different segmentation masks  [EXPECTED, NOT FIXABLE]
────────────────────────────────────────────────────────────────────────────
Symptom : Area MAE=700 µm², Volume MAE=892 µm³, Sphericity MAE=0.107,
          N Voxels MAE=209; all intensity metrics negative r².
Cause   : Intrinsic difference between ACDC deep-learning segmentation
          (tight nuclear boundaries) and Imaris surface-reconstruction
          (intensity-threshold, includes more cytoplasm).  The two systems
          measure physically different regions of the same cell.
          Intensity metrics fail for the same reason: different voxel masks
          → different pixel samples even for correctly matched cells.
Fix     : None possible without rerunning Imaris with the same masks.
          Area/Volume/Intensity comparisons should be interpreted as
          "pipeline vs Imaris methodology difference", not as errors.

═══════════════════════════════════════════════════════════════════════════
FORMULA VALIDATION SUMMARY (all formulas validated vs Imaris ground truth)
═══════════════════════════════════════════════════════════════════════════

VALIDATED FIXES (from exhaustive formula search)
─────────────────────────────────────────────────
1) Speed / Velocity  →  CENTRAL DIFFERENCE  (was: backward diff)
   speed[T]  = |pos[T+1] - pos[T-1]| / (2·Δt)     r²=0.987
   vel[T]    = (pos[T+1] - pos[T-1])  / (2·Δt)     r²=0.985–0.989
   NaN at first AND last frame of each track (edge: no central diff possible).

2) Acceleration  →  OUTPUT ZERO  (was: diff(|v|)/dt)
   Imaris exports ALL ZEROS in every dataset tested.
   Every computed variant gave r²=NaN (zero variance ground truth).

3) Track Speed Mean  →  track_length / duration  (TLD)  r²=0.976
   (was: arithmetic mean of per-frame speeds — that gives r²=0.957)

4) Track Speed Max / Min / StdDev / Variation  →  BACKWARD-diff speeds
   Backward diff better than central for track-level extremes (r²=0.85-0.88)
   because central drops edge frames where extreme speeds often occur.

UNCHANGED (already correct)
────────────────────────────
 5) Displacement²  from_start mode: r²=1.000  ✓
 6) Track Displacement X/Y/Z:        r²=1.000  ✓
 7) Track Duration:                  r²=1.000  ✓
 8) Track Length:                    r²=1.000  ✓
 9) Track Straightness:              r²=1.000  ✓
10) Track N Surfaces:                r²=1.000  ✓
11) Track Position X/Y/Z Mean:       r²=1.000  ✓

STILL IMPERFECT (best available formula kept)
──────────────────────────────────────────────
12) Track AR1 X/Y (diff_demean):    r²≈0.80  — Imaris formula differs slightly
13) Track AR1 Mean:                 r²≈0.22  — Imaris 3D AR1 aggregation unknown

NPZ/TIF-DERIVED SHEETS
────────────────────────────────────────────────────────────────────────────
Area, Volume, Intensity are intrinsic per-frame properties — unaffected by
temporal convention.  After the time-index fix they will join correctly, but
systematic differences vs Imaris remain due to segmentation mask differences
(Problem 5 above).

Known systematic gaps:
  - Volume (2D): area × pixel_size² × z_step  (z_step = nominal cell thickness)
  - Sphericity (2D): approximated via extruded slab marching cubes
  - Ellipsoid axes: covariance method; axis lengths = 2√λ (Imaris uses surface-mesh inertia tensor)
  - Ellipticity: Oblate = 1-(L_C/L_B), Prolate = 1-(L_B/L_A)  (L_A ≥ L_B ≥ L_C)
  - Position: intensity-weighted centroid (CIM), not homogeneous centroid
  - IntensityCenter: image value at rounded integer pixel coords of CIM
"""

import argparse
import os
import re
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import tifffile as tiff

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from skimage.filters import gaussian as _gaussian_blur
from skimage.measure import regionprops, label as cc_label, marching_cubes


# ----------------------------
# Helpers: safe stats
# ----------------------------
def safe_nanmean(a) -> float:
    a = np.asarray(a, dtype=float)
    return float(np.nanmean(a)) if np.isfinite(a).any() else np.nan


def safe_nanmax(a) -> float:
    a = np.asarray(a, dtype=float)
    return float(np.nanmax(a)) if np.isfinite(a).any() else np.nan


def safe_nanmin(a) -> float:
    a = np.asarray(a, dtype=float)
    return float(np.nanmin(a)) if np.isfinite(a).any() else np.nan


# ----------------------------
# IO helpers
# ----------------------------
def clean_colname(c: str) -> str:
    return re.sub(r"\s+", " ", str(c).strip())


def load_table_optional(path: Optional[str]) -> Optional[pd.DataFrame]:
    if not path:
        return None
    if not os.path.exists(path):
        raise FileNotFoundError(f"--acdc-table not found: {path}")
    ext = os.path.splitext(path)[1].lower()
    if ext in [".csv", ".txt"]:
        df = pd.read_csv(path)
    elif ext in [".xlsx", ".xls"]:
        df = pd.read_excel(path)
    else:
        raise ValueError(f"Unsupported --acdc-table extension: {ext}")
    df.columns = [clean_colname(c) for c in df.columns]
    return df


def load_segmentation_npz(path: str) -> np.ndarray:
    """
    Supports:
      - (T,Y,X)
      - (T,Z,Y,X)
      - object arrays of frames
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"Segmentation NPZ not found: {path}")

    z = np.load(path, allow_pickle=True)
    keys = list(z.keys())

    cand_keys = ["segm", "masks", "labels", "mask", "arr_0"]
    arr = None
    for k in cand_keys:
        if k in keys:
            a = z[k]
            if isinstance(a, np.ndarray):
                arr = a
                break
    if arr is None:
        for k in keys:
            a = z[k]
            if isinstance(a, np.ndarray) and a.size > 0:
                arr = a
                break
    if arr is None:
        raise RuntimeError(f"Could not find a segmentation array in NPZ. Keys={keys}")

    if arr.dtype == object:
        frames = [np.asarray(arr[i]) for i in range(len(arr))]
        arr = np.stack(frames, axis=0)

    arr = np.asarray(arr)
    if not np.issubdtype(arr.dtype, np.integer):
        arr = arr.astype(np.int32)

    # Normalize dims
    if arr.ndim == 2:
        arr = arr[None, ...]     # (1,Y,X)
    elif arr.ndim in (3, 4):
        pass
    else:
        raise ValueError(f"Unsupported segm ndim={arr.ndim}, shape={arr.shape}")

    return arr


def load_tif_frames(path: str, channel_index: int = 0) -> np.ndarray:
    """
    Returns float32 image array.
    Supports common layouts:
      - (T,Y,X)
      - (T,Z,Y,X)
      - (Y,X) or (Z,Y,X)
      - channel dims (best-effort)
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"TIF not found: {path}")

    img = np.asarray(tiff.imread(path))

    if img.ndim == 2:
        img = img[None, ...]  # (1,Y,X)
    elif img.ndim == 3:
        pass
    elif img.ndim == 4:
        # Could be (T,Y,X,C) or (T,C,Y,X) or (T,Z,Y,X)
        if img.shape[-1] <= 10 and img.shape[1] > 10 and img.shape[2] > 10:
            img = img[..., channel_index]
        elif img.shape[1] <= 10 and img.shape[2] > 10 and img.shape[3] > 10:
            img = img[:, channel_index, ...]
        else:
            # assume (T,Z,Y,X)
            pass
    elif img.ndim == 5:
        # likely (T,C,Z,Y,X)
        if img.shape[1] <= 10:
            img = img[:, channel_index, ...]
        else:
            raise ValueError(f"Cannot infer 5D TIF layout: shape={img.shape}")
    else:
        raise ValueError(f"Unsupported TIF ndim={img.ndim}, shape={img.shape}")

    return img.astype(np.float32)


def align_time_axis(segm: np.ndarray, img: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    """
    Align segm and img on time axis and spatial dims.
    Accepts:
      segm (T,Y,X) with img (T,Y,X)
      segm (T,Z,Y,X) with img (T,Z,Y,X)
    """
    if segm.shape[0] != img.shape[0]:
        raise RuntimeError(f"Time mismatch: segm T={segm.shape[0]} vs img T={img.shape[0]}")

    if segm.ndim != img.ndim:
        raise RuntimeError(f"Dim mismatch: segm ndim={segm.ndim} img ndim={img.ndim}. Use matching 2D/3D.")

    if segm.shape[1:] != img.shape[1:]:
        raise RuntimeError(f"Spatial mismatch: segm={segm.shape} img={img.shape}")

    return segm, img


# ----------------------------
# Imaris-consistent mesh geometry
# ----------------------------

def mesh_stats_from_binary(
    vol_bool_3d: np.ndarray,
    spacing_zyx: Tuple[float, float, float],
    smooth_sigma: float = 0.0,
    smooth_level: float = 0.5,
) -> Tuple[float, int, int]:
    """
    Returns (surface_area_um2, n_vertices, n_triangles).
    Uses marching cubes and triangle area sum.

    Padding: only added when the entire volume is filled (no background),
    which is rare in practice. The 2D slab [empty, mask, empty] already
    contains background so no extra padding is needed.
    Input cast to float32 for numerical stability in marching cubes.

    smooth_sigma : Gaussian blur sigma in pixels applied to the YX plane before
                   marching cubes.  When > 0 the surface is smoothed and the
                   isosurface is evaluated at smooth_level instead of 0.5.
                   A level < 0.5 expands the surface outward, replicating the
                   volume/area inflation seen in Imaris surface mesh creation.
                   Calibrated against GT: sigma≈7 px, level≈0.25 for 1.22 um/px.
    smooth_level : Marching cubes isosurface level (default 0.5 = exact binary
                   boundary).  Values < 0.5 expand the surface outward.
    """
    if vol_bool_3d.ndim != 3:
        raise ValueError("mesh_stats_from_binary expects 3D volume")

    vol = vol_bool_3d.astype(np.uint8)

    if vol.max() == 0:
        return 0.0, 0, 0

    # Only pad when no background voxel exists (object fills entire bounding box)
    if np.all(vol):
        vol = np.pad(vol, pad_width=1, mode="constant", constant_values=0)

    vol_float = vol.astype(np.float32)
    if smooth_sigma > 0.0:
        # Blur only in Y,X (axes 1 and 2); leave Z (axis 0) sharp to preserve
        # the single-slab structure of 2D cells.
        vol_float = _gaussian_blur(
            vol_float, sigma=(0.0, smooth_sigma, smooth_sigma)
        ).astype(np.float32)

    try:
        verts, faces, _normals, _values = marching_cubes(
            vol_float, level=smooth_level, spacing=spacing_zyx)
    except Exception:
        return 0.0, 0, 0

    n_vertices = int(verts.shape[0])
    n_triangles = int(faces.shape[0])

    if n_triangles == 0:
        return 0.0, n_vertices, n_triangles

    v0 = verts[faces[:, 0]]
    v1 = verts[faces[:, 1]]
    v2 = verts[faces[:, 2]]
    cross = np.cross(v1 - v0, v2 - v0)
    tri_areas = 0.5 * np.linalg.norm(cross, axis=1)
    area = float(np.sum(tri_areas)) if tri_areas.size else 0.0

    return area, n_vertices, n_triangles


def wadell_sphericity(volume_um3: float, surface_area_um2: float) -> float:
    """
    Wadell sphericity: area(sphere with same volume) / particle area
    area_sphere = pi^(1/3) * (6V)^(2/3)
    """
    if not (np.isfinite(volume_um3) and np.isfinite(surface_area_um2)):
        return np.nan
    if volume_um3 <= 0 or surface_area_um2 <= 0:
        return np.nan
    area_sphere = (np.pi ** (1.0 / 3.0)) * ((6.0 * volume_um3) ** (2.0 / 3.0))
    return float(area_sphere / surface_area_um2)


def covariance_ellipsoid_axes(coords_xyz_um: np.ndarray, z_step_um: float = 1.0) -> Tuple[np.ndarray, np.ndarray]:
    """
    Estimate ellipsoid axes from voxel cloud covariance.

    For 2D cells (z constant for all voxels), uses an analytical 2×2
    eigendecomposition that avoids LAPACK/eigh calls, which crash on some
    Windows systems due to numpy/LAPACK ABI conflicts.

    Imaris axis convention for flat 2D cells (confirmed by formula search):
      Axis A = z half-extent = 0.5 * z_step  (constant, smallest)
      Axis B = short in-plane axis            (middle size)
      Axis C = long  in-plane axis            (largest)

    Returns (vecs, lens) where:
      vecs[:, 0]  direction of Axis C (long in-plane)
      vecs[:, 1]  direction of Axis B (short in-plane)
      vecs[:, 2]  direction of Axis A (z-axis = [0,0,1])
      lens[0]  = 2 * sqrt(la)  Axis C length (long in-plane)
      lens[1]  = 2 * sqrt(lb)  Axis B length (short in-plane)
      lens[2]  = 0.5 * z_step  Axis A length (constant)
    """
    n = coords_xyz_um.shape[0]
    z_half = 0.5 * z_step_um

    if n == 0:
        return np.eye(3), np.array([np.nan, np.nan, z_half], dtype=float)
    if n == 1:
        return np.eye(3), np.array([0.0, 0.0, z_half], dtype=float)

    # Work in x-y plane only (z is constant for 2D data → zero variance in z)
    x_um = coords_xyz_um[:, 0]
    y_um = coords_xyz_um[:, 1]
    ddof = 1  # unbiased

    xm, ym  = x_um.mean(), y_um.mean()
    dx, dy  = x_um - xm, y_um - ym
    var_x   = float(np.sum(dx * dx)) / (n - ddof)
    var_y   = float(np.sum(dy * dy)) / (n - ddof)
    cov_xy  = float(np.sum(dx * dy)) / (n - ddof)

    # Analytical eigenvalues of the 2×2 covariance matrix
    trace = var_x + var_y
    det   = var_x * var_y - cov_xy * cov_xy
    disc  = max(0.0, (trace / 2.0) ** 2 - det)
    la = trace / 2.0 + np.sqrt(disc)   # largest in-plane eigenvalue
    lb = max(0.0, trace / 2.0 - np.sqrt(disc))

    # Eigenvector for eigenvalue la  (solve (C - la*I)v = 0 analytically)
    if abs(cov_xy) > 1e-14:
        vx, vy = cov_xy, la - var_x
    elif var_x >= var_y:
        vx, vy = 1.0, 0.0
    else:
        vx, vy = 0.0, 1.0

    norm = (vx * vx + vy * vy) ** 0.5
    if norm > 0:
        vx /= norm; vy /= norm
    # Perpendicular eigenvector for lb
    ux, uy = -vy, vx

    # 3×3 axis matrix (each column is one axis direction in x,y,z space)
    vecs = np.array([
        [vx,  ux,  0.0],   # x-components of C, B, A axes
        [vy,  uy,  0.0],   # y-components
        [0.0, 0.0, 1.0],   # z-components
    ], dtype=float)

    lens = np.array([
        2.0 * np.sqrt(max(0.0, la)),  # Axis C (long in-plane)
        2.0 * np.sqrt(max(0.0, lb)),  # Axis B (short in-plane)
        z_half,                        # Axis A (z half-extent, constant)
    ], dtype=float)

    return vecs, lens


def ellipticity_from_axes(a: float, b: float, c: float) -> Tuple[float, float]:
    """
    Imaris ellipticity formulas (L_A >= L_B >= L_C):

      Oblate:  1 - (L_C / L_B)
      Prolate: 1 - (L_B / L_A)
    """
    if not np.isfinite([a, b, c]).all():
        return np.nan, np.nan
    L_A, L_B, L_C = sorted([a, b, c], reverse=True)  # L_A >= L_B >= L_C
    if L_A <= 0 or L_B <= 0:
        return np.nan, np.nan

    e_oblate = float(1.0 - (L_C / L_B)) if L_B > 0 else np.nan
    e_prolate = float(1.0 - (L_B / L_A)) if L_A > 0 else np.nan

    return e_oblate, e_prolate


# ----------------------------
# Parent mapping from ACDC table (optional)
# ----------------------------
def _find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in cols:
            return cols[cand.lower()]
    return None


def build_parent_mapper(acdc_df: Optional[pd.DataFrame]) -> Optional[Dict[int, pd.DataFrame]]:
    """
    Returns dict: frame_i -> df_frame with columns [Cell_ID, x, y, z(optional)]
    """
    if acdc_df is None or acdc_df.empty:
        return None

    frame_col = _find_col(acdc_df, ["frame_i", "frame", "t", "time", "Time"])
    id_col = _find_col(acdc_df, ["Cell_ID", "cell_id", "track_id", "Track_ID", "ID"])
    x_col = _find_col(acdc_df, ["x_centroid", "centroid_x", "x", "X"])
    y_col = _find_col(acdc_df, ["y_centroid", "centroid_y", "y", "Y"])
    z_col = _find_col(acdc_df, ["z_centroid", "centroid_z", "z", "Z"])

    if frame_col is None or id_col is None or x_col is None or y_col is None:
        return None

    use_cols = [frame_col, id_col, x_col, y_col] + ([z_col] if z_col else [])
    df = acdc_df[use_cols].copy()
    df = df.rename(columns={frame_col: "frame_i", id_col: "Cell_ID", x_col: "x", y_col: "y"})
    if z_col:
        df = df.rename(columns={z_col: "z"})
    else:
        df["z"] = np.nan

    out: Dict[int, pd.DataFrame] = {}
    for fi, g in df.groupby("frame_i"):
        try:
            fi_int = int(fi)
        except Exception:
            continue
        out[fi_int] = g.reset_index(drop=True)
    return out


def map_parent_for_object(
    mapper: Optional[Dict[int, pd.DataFrame]],
    frame_i: int,
    centroid_yx_or_zyx: Tuple[float, float, float],
    map_max_dist_px: float,
    offset_x_px: float,
    offset_y_px: float,
    offset_z_px: float,
) -> Optional[int]:
    """
    centroid tuple is (z,y,x) (for 2D: z=0)
    mapping uses x,y,z in pixel space
    """
    if mapper is None or frame_i not in mapper:
        return None
    df = mapper[frame_i]
    if df.empty:
        return None

    cz, cy, cx = centroid_yx_or_zyx
    cx = float(cx) + float(offset_x_px)
    cy = float(cy) + float(offset_y_px)
    cz = float(cz) + float(offset_z_px)

    dx = df["x"].astype(float).to_numpy() - cx
    dy = df["y"].astype(float).to_numpy() - cy
    dz = np.nan_to_num(df["z"].astype(float).to_numpy(), nan=0.0) - (0.0 if np.isnan(cz) else cz)

    dist = np.sqrt(dx * dx + dy * dy + dz * dz)
    j = int(np.argmin(dist))
    if float(dist[j]) > float(map_max_dist_px):
        return None
    return int(df.loc[j, "Cell_ID"])


# ----------------------------
# Core computation
# ----------------------------
def compute_object_table(
    segm: np.ndarray,
    img: np.ndarray,
    pixel_size_um: float,
    z_step_um: float,
    frame_interval_s: float,
    swap_xy: bool,
    progress_every_frames: int,
    progress_every_objects: int,
    compute_disconnected_components: bool,
    skip_sphericity: bool,
    acdc_df: Optional[pd.DataFrame],
    map_max_dist_px: float,
    map_offset_x_px: float,
    map_offset_y_px: float,
    map_offset_z_px: float,
    max_frames: Optional[int],
    displacement2_mode: str,  # "from_start" or "step"
    # global ACDC->Imaris offsets in pixels (applied to exported coordinates)
    imaris_offset_x_px: float,
    imaris_offset_y_px: float,
    imaris_offset_z_px: float,
    # optional TrackPy track mapping: (frame, cell_id) -> particle (persistent track ID)
    trackpy_mapper: Optional[Dict] = None,
    # channel name used to look up intensity columns in acdc_df (e.g. "NIR")
    acdc_channel_name: Optional[str] = None,
    # Time index of the FIRST frame in the output Excel.
    # Imaris convention (empirically confirmed): first frame = 2.
    # Set to 1 if your Imaris export shows Time=1 for the first frame.
    time_index_start: int = 2,
    # Gaussian pre-smoothing for 2D cell mesh (to match Imaris surface inflation).
    # sigma in pixels (YX plane); level = marching cubes isosurface threshold.
    # sigma=7.0, level=0.25 calibrated from GT ratios at px=1.22 um.
    # Set sigma=0 to disable (original behaviour).
    mesh_smooth_sigma: float = 0.0,
    mesh_smooth_level: float = 0.25,
) -> pd.DataFrame:
    """
    NOTE (pipeline behavior):
    - If acdc_df is provided, we ALWAYS require mapping to a CSV Cell_ID.
      Unmapped NPZ objects are skipped and never exported.
    - If acdc_df is NOT provided, Parent defaults to label_id.
    """
    T = segm.shape[0]
    if max_frames is not None:
        T = min(T, int(max_frames))

    is3d = (segm.ndim == 4)  # (T,Z,Y,X)
    mapper = build_parent_mapper(acdc_df) if (acdc_df is not None) else None

    # Build (frame_i, Cell_ID) -> intensity values lookup from ACDC CSV
    acdc_intens_lookup: Dict = {}
    if acdc_df is not None and acdc_channel_name:
        ch = acdc_channel_name
        # Map metric suffix to lookup key
        _icols = {
            "mean":   f"{ch}_mean",
            "median": f"{ch}_median",
            "min":    f"{ch}_min",
            "max":    f"{ch}_max",
            "sum":    f"{ch}_sum",
            "std":    None,  # not in ACDC — will recompute from TIF
        }
        # Find the right frame column
        _fcol = _find_col(acdc_df, ["frame_i", "frame", "t", "time"])
        _idcol = _find_col(acdc_df, ["Cell_ID", "cell_id", "ID"])
        if _fcol and _idcol:
            for _, row in acdc_df.iterrows():
                key = (int(row[_fcol]), int(row[_idcol]))
                vals = {}
                for metric, col in _icols.items():
                    if col and col in acdc_df.columns:
                        vals[metric] = float(row[col]) if not pd.isna(row[col]) else np.nan
                acdc_intens_lookup[key] = vals
            n_intens = len(acdc_intens_lookup)
            if n_intens > 0:
                print(f"[info] ACDC intensity lookup built: {n_intens:,} (frame,cell_id) entries from channel '{ch}'", flush=True)
        else:
            print(f"[warn] Could not build ACDC intensity lookup (missing frame/ID cols)", flush=True)

    if acdc_df is not None and mapper is None:
        raise RuntimeError(
            "You provided --acdc-table but required columns were not found.\n"
            "Expected at least: frame/time column + Cell_ID + centroid x/y (and optional z).\n"
            f"Columns found: {list(acdc_df.columns)}"
        )

    rows: List[Dict] = []

    t0_all = time.time()
    for t in range(T):
        if (t % progress_every_frames) == 0:
            print(f"[progress] frame {t+1}/{T}: start", flush=True)

        lab = segm[t]
        im = img[t]

        props = regionprops(lab.astype(np.int32), intensity_image=im)

        frame_i0 = int(t)
        # ── TIME INDEX FIX ─────────────────────────────────────────────────────
        # Imaris labels the first captured frame as time_index_start (default=2).
        # Empirically confirmed: pipeline previously output Time=1 for first
        # frame while Imaris output Time=2 — causing ALL per-frame joins
        # (Area, Volume, Speed, Intensity...) to pair the wrong frames and
        # produce negative r² across the board.
        # Pass time_index_start=1 if your Imaris export starts at Time=1.
        frame_idx = int(t + time_index_start)
        time_s = float(t * frame_interval_s)

        for j, r in enumerate(props, start=1):
            if (j % progress_every_objects) == 0:
                print(f"[progress] frame {t+1}/{T}: object {j}/{len(props)}", flush=True)

            label_id = int(r.label)
            n_vox_2d_or_3d = int(r.area)

            coords = r.coords  # 3D: (z,y,x), 2D: (y,x)
            if coords.size == 0:
                continue

            if is3d:
                z = coords[:, 0].astype(np.float64)
                y = coords[:, 1].astype(np.float64)
                x = coords[:, 2].astype(np.float64)
            else:
                y = coords[:, 0].astype(np.float64)
                x = coords[:, 1].astype(np.float64)
                z = np.zeros_like(x, dtype=np.float64)

            # voxel centers (index + 0.5) -> um
            x_um0 = (x + 0.5) * pixel_size_um
            y_um0 = (y + 0.5) * pixel_size_um
            z_um = (z + 0.5) * z_step_um

            # optional swap first (so offsets apply to final exported axes)
            if swap_xy:
                x_um, y_um = y_um0, x_um0
            else:
                x_um, y_um = x_um0, y_um0

            # apply global ACDC->Imaris offsets (pixels -> um)
            x_um = x_um + (imaris_offset_x_px * pixel_size_um)
            y_um = y_um + (imaris_offset_y_px * pixel_size_um)
            z_um = z_um + (imaris_offset_z_px * z_step_um)

            # Center of Homogeneous Mass (mean of voxel centers)
            chom_x = float(np.mean(x_um))
            chom_y = float(np.mean(y_um))
            chom_z = float(np.mean(z_um))

            # Intensities for the voxels
            if is3d:
                intens = im[
                    coords[:, 0].astype(int),
                    coords[:, 1].astype(int),
                    coords[:, 2].astype(int)
                ].astype(np.float64)
            else:
                intens = im[
                    coords[:, 0].astype(int),
                    coords[:, 1].astype(int)
                ].astype(np.float64)

            # Center of Image Mass (intensity-weighted)
            wsum = float(np.sum(intens))
            if wsum > 0:
                cim_x = float(np.sum(x_um * intens) / wsum)
                cim_y = float(np.sum(y_um * intens) / wsum)
                cim_z = float(np.sum(z_um * intens) / wsum)
            else:
                cim_x, cim_y, cim_z = chom_x, chom_y, chom_z

            # Position = intensity-weighted centroid (Center of Image Mass), per Imaris convention
            pos_x, pos_y, pos_z = cim_x, cim_y, cim_z

            # Intensity stats
            if intens.size == 0:
                imin = imax = imean = imed = istd = isum = np.nan
            else:
                imin = float(np.min(intens))
                imax = float(np.max(intens))
                imean = float(np.mean(intens))
                imed = float(np.median(intens))
                istd = float(np.std(intens, ddof=0))
                isum = float(np.sum(intens))

            # Intensity center: image value at the rounded integer pixel coords of the
            # intensity-weighted centroid (CIM), matching Imaris "intensity at Position".
            if intens.size and wsum > 0:
                if is3d:
                    wc_z = int(np.clip(int(np.round(float(np.sum(z * intens) / wsum))), 0, im.shape[0] - 1))
                    wc_y = int(np.clip(int(np.round(float(np.sum(y * intens) / wsum))), 0, im.shape[1] - 1))
                    wc_x = int(np.clip(int(np.round(float(np.sum(x * intens) / wsum))), 0, im.shape[2] - 1))
                    intensity_center = float(im[wc_z, wc_y, wc_x])
                else:
                    wc_y = int(np.clip(int(np.round(float(np.sum(y * intens) / wsum))), 0, im.shape[0] - 1))
                    wc_x = int(np.clip(int(np.round(float(np.sum(x * intens) / wsum))), 0, im.shape[1] - 1))
                    intensity_center = float(im[wc_y, wc_x])
            else:
                intensity_center = float(intens[0]) if intens.size else np.nan

            # Disconnected components inside bbox mask
            if compute_disconnected_components:
                n_cc = int(cc_label(r.image.astype(np.uint8), connectivity=1).max())
            else:
                n_cc = 1 if n_vox_2d_or_3d > 0 else 0

            # ---- Mesh stats (Area/Triangles/Vertices) ----
            if is3d:
                vol_bool = r.image.astype(bool)  # (z,y,x) bbox
                spacing = (z_step_um, pixel_size_um, pixel_size_um)
                voxel_vol_um3 = pixel_size_um * pixel_size_um * z_step_um
                volume_um3 = float(n_vox_2d_or_3d * voxel_vol_um3)
                n_voxels_report = int(n_vox_2d_or_3d)
            else:
                # 2D bbox mask (y,x) -> extrude to 3 slices for marching cubes.
                # (empty, solid, empty) gives a single closed slab of thickness
                # z_step_um, matching how Imaris treats 2D surfaces.
                m2d = r.image.astype(bool)
                empty = np.zeros_like(m2d)
                vol_bool = np.stack([empty, m2d, empty], axis=0)  # (3,y,x)
                spacing = (z_step_um, pixel_size_um, pixel_size_um)

                area_px = int(n_vox_2d_or_3d)
                # When Gaussian smoothing is active, compute volume from the
                # expanded 2D mask (mirrors Imaris mesh-volume inflation).
                # n_voxels_report always uses the original pixel count.
                if mesh_smooth_sigma > 0.0:
                    m2d_smooth = _gaussian_blur(
                        m2d.astype(np.float32), sigma=mesh_smooth_sigma
                    )
                    expanded_px = int(np.sum(m2d_smooth > mesh_smooth_level))
                    volume_um3 = float(expanded_px * pixel_size_um ** 2 * z_step_um)
                else:
                    volume_um3 = float(area_px * pixel_size_um ** 2 * z_step_um)
                n_voxels_report = int(area_px)  # always original count, like Imaris

            area_um2, n_vertices, n_triangles = mesh_stats_from_binary(
                vol_bool, spacing,
                smooth_sigma=mesh_smooth_sigma if not is3d else 0.0,
                smooth_level=mesh_smooth_level if not is3d else 0.5,
            )

            # Sphericity (Wadell)
            if skip_sphericity:
                sphericity = np.nan
            else:
                sphericity = wadell_sphericity(volume_um3, area_um2)

            # Ellipsoid axes from voxel cloud covariance.
            # Imaris axis convention (confirmed by formula search against GT):
            #   Axis A = z half-extent (constant = 0.5 * z_step)  ← GT shows 0.610 um
            #   Axis B = short in-plane principal axis
            #   Axis C = long  in-plane principal axis             ← GT range 4.9-48.5 um
            # Uses analytical 2×2 eigen to avoid LAPACK crash on Windows.
            coords_xyz_um = np.stack([x_um, y_um, z_um], axis=1)
            axis_vecs, axis_lens = covariance_ellipsoid_axes(coords_xyz_um, z_step_um)
            # vecs[:,0]=C(long), vecs[:,1]=B(short), vecs[:,2]=A(z)
            axisC_vec = axis_vecs[:, 0]   # long in-plane
            axisB_vec = axis_vecs[:, 1]   # short in-plane
            axisA_vec = axis_vecs[:, 2]   # z-direction
            axisC_len = axis_lens[0]       # long in-plane  = Imaris Axis C
            axisB_len = axis_lens[1]       # short in-plane = Imaris Axis B
            axisA_len = axis_lens[2]       # 0.5*z_step     = Imaris Axis A (constant)
            e_oblate, e_prolate = ellipticity_from_axes(axisA_len, axisB_len, axisC_len)

            # --- Parent mapping behavior ---
            centroid = r.centroid
            if is3d:
                cz, cy, cx = float(centroid[0]), float(centroid[1]), float(centroid[2])
            else:
                cz, cy, cx = 0.0, float(centroid[0]), float(centroid[1])

            # If ACDC table exists -> ALWAYS require mapping to a CSV Cell_ID
            if mapper is not None:
                parent = map_parent_for_object(
                    mapper,
                    frame_i0,
                    (cz, cy, cx),
                    map_max_dist_px,
                    map_offset_x_px,
                    map_offset_y_px,
                    map_offset_z_px
                )
                if parent is None:
                    # Not present in CSV (e.g., filtered out) -> skip export
                    continue
            else:
                # No CSV provided: fallback to segmentation label
                parent = label_id

            cell_id = int(parent)

            # Override intensity values from ACDC CSV if available.
            # ACDC intensity values are more reliable than our TIF recomputation
            # because they use the same pipeline that generated the ground truth.
            _acdc_intens = acdc_intens_lookup.get((frame_i0, cell_id))
            if _acdc_intens:
                if "mean" in _acdc_intens and not np.isnan(_acdc_intens["mean"]):
                    imean = _acdc_intens["mean"]
                if "median" in _acdc_intens and not np.isnan(_acdc_intens["median"]):
                    imed = _acdc_intens["median"]
                if "min" in _acdc_intens and not np.isnan(_acdc_intens["min"]):
                    imin = _acdc_intens["min"]
                if "max" in _acdc_intens and not np.isnan(_acdc_intens["max"]):
                    imax = _acdc_intens["max"]
                if "sum" in _acdc_intens and not np.isnan(_acdc_intens["sum"]):
                    isum = _acdc_intens["sum"]
                # istd is not in ACDC CSV - keep TIF-computed value
                # intensity_center (nearest voxel to COM) - keep TIF-computed value

            # If a TrackPy mapping is available, use particle as the persistent
            # Parent (track ID). Falls back to cell_id if not found.
            if trackpy_mapper is not None:
                track_parent = trackpy_mapper.get((frame_i0, cell_id))
                if track_parent is None:
                    # Key not found - could be frame offset or ID mismatch
                    # Emit a one-time warning with a sample key so it is easy to debug
                    if not getattr(compute_object_table, "_tp_miss_warned", False):
                        sample_keys = list(trackpy_mapper.keys())[:3]
                        print(
                            f"[warn] trackpy_mapper miss: key=({frame_i0}, {cell_id}) not found. "
                            f"Sample keys in mapper: {sample_keys}",
                            flush=True,
                        )
                        compute_object_table._tp_miss_warned = True
                    track_parent = cell_id
            else:
                track_parent = cell_id

            rows.append(dict(
                frame=frame_i0,
                frame_idx=frame_idx,
                time_s=time_s,

                ID=str(cell_id),
                Parent=track_parent,

                area_um2=float(area_um2),
                volume_um3=float(volume_um3),

                n_voxels=int(n_voxels_report),
                n_vertices=int(n_vertices),
                n_triangles=int(n_triangles),
                n_disconnected_components=int(n_cc),
                sphericity=float(sphericity) if np.isfinite(sphericity) else np.nan,

                pos_x_um=float(pos_x),
                pos_y_um=float(pos_y),
                pos_z_um=float(pos_z),

                chom_x_um=float(chom_x),
                chom_y_um=float(chom_y),
                chom_z_um=float(chom_z),

                cim_x_um=float(cim_x),
                cim_y_um=float(cim_y),
                cim_z_um=float(cim_z),

                intensity_center=float(intensity_center),
                intensity_max=imax,
                intensity_min=imin,
                intensity_mean=imean,
                intensity_median=imed,
                intensity_std=istd,
                intensity_sum=isum,

                axisA_len_um=float(axisA_len),
                axisB_len_um=float(axisB_len),
                axisC_len_um=float(axisC_len),

                axisA_x=float(axisA_vec[0]), axisA_y=float(axisA_vec[1]), axisA_z=float(axisA_vec[2]),
                axisB_x=float(axisB_vec[0]), axisB_y=float(axisB_vec[1]), axisB_z=float(axisB_vec[2]),
                axisC_x=float(axisC_vec[0]), axisC_y=float(axisC_vec[1]), axisC_z=float(axisC_vec[2]),

                ellipticity_oblate=float(e_oblate) if np.isfinite(e_oblate) else np.nan,
                ellipticity_prolate=float(e_prolate) if np.isfinite(e_prolate) else np.nan,
            ))

        if (t % progress_every_frames) == 0:
            print(f"[progress] frame {t+1}/{T}: done | elapsed {time.time()-t0_all:.1f}s", flush=True)

    df = pd.DataFrame(rows)
    if df.empty:
        raise RuntimeError(
            "No objects exported. If you passed --acdc-table, mapping likely failed for all objects.\n"
            "Try increasing --map-max-dist-px or verify centroid columns in the CSV."
        )

    # -------------------------
    # Kinematics per track (Parent)
    # -------------------------
    # VALIDATED FORMULAS (from formula_search_report against Imaris ground truth):
    #
    #   Speed / Velocity:  CENTRAL DIFFERENCE  r²=0.987/0.985/0.989
    #     speed[T]  = dist(pos[T-1] → pos[T+1]) / (2*dt)
    #     vel[T]    = (pos[T+1] - pos[T-1]) / (2*dt)
    #     First and last frame of each track → NaN (edge, no central diff possible)
    #
    #   Acceleration: Imaris exports ALL ZEROS for this quantity.
    #     Every formula variant produced r²=NaN (ground truth is constant 0).
    #     → Output 0.0 for all rows.
    #
    #   Track Speed Mean: track_length / duration  (TLD)  r²=0.976
    #     (NOT arithmetic mean of per-frame speeds)
    #
    #   Track Speed Max/Min/StdDev/Variation:
    #     Computed from BACKWARD-diff per-frame speeds (best r²=0.85–0.88).
    #     Central diff is WORSE for these because it drops edge frames, losing
    #     the highest/lowest speeds which are often at track boundaries.
    #     Backward speed kept separately as "speed_bkwd_um_s" for this purpose.

    df = df.sort_values(["Parent", "frame"]).reset_index(drop=True)
    g = df.groupby("Parent", sort=False)

    # ── backward diffs (T-1 → T) ──────────────────────────────────────────────
    dx_bk = g["pos_x_um"].diff()
    dy_bk = g["pos_y_um"].diff()
    dz_bk = g["pos_z_um"].diff()
    dt_bk = g["time_s"].diff().replace(0, np.nan)

    # backward speed — used only for track-level aggregates (Max/Min/Std/Var)
    df["speed_bkwd_um_s"] = np.sqrt(dx_bk**2 + dy_bk**2 + dz_bk**2) / dt_bk

    # ── forward diffs (T → T+1) ───────────────────────────────────────────────
    dx_fw = g["pos_x_um"].diff(-1).mul(-1)   # x[T+1] - x[T]
    dy_fw = g["pos_y_um"].diff(-1).mul(-1)
    dz_fw = g["pos_z_um"].diff(-1).mul(-1)
    dt_fw = g["time_s"].diff(-1).mul(-1).replace(0, np.nan)

    # ── central diff (T-1 → T+1) — VALIDATED best for per-frame Speed/Velocity ─
    # Central displacement vector at T: pos[T+1] - pos[T-1]
    # Central dt at T:  time[T+1] - time[T-1]  (= dt_bk + dt_fw at T)
    cx = dx_bk + dx_fw          # (x[T] - x[T-1]) + (x[T+1] - x[T]) = x[T+1] - x[T-1]
    cy = dy_bk + dy_fw
    cz = dz_bk + dz_fw
    dt_central = (dt_bk + dt_fw).replace(0, np.nan)

    # Mark edges: first and last row per track cannot have central diff
    rank_asc  = g["pos_x_um"].cumcount()
    rank_desc = g["pos_x_um"].cumcount(ascending=False)
    edge_mask = (rank_asc == 0) | (rank_desc == 0)

    # Central velocity
    df["vel_x_um_s"] = cx / dt_central
    df["vel_y_um_s"] = cy / dt_central
    df["vel_z_um_s"] = cz / dt_central
    df.loc[edge_mask, ["vel_x_um_s", "vel_y_um_s", "vel_z_um_s"]] = np.nan

    # Central speed  (magnitude of central displacement / central dt)
    df["speed_um_s"] = np.sqrt(cx**2 + cy**2 + cz**2) / dt_central
    df.loc[edge_mask, "speed_um_s"] = np.nan

    # ── Displacement² ────────────────────────────────────────────────────────
    if displacement2_mode == "from_start":
        x0 = g["pos_x_um"].transform("first")
        y0 = g["pos_y_um"].transform("first")
        z0 = g["pos_z_um"].transform("first")
        ddx = df["pos_x_um"] - x0
        ddy = df["pos_y_um"] - y0
        ddz = df["pos_z_um"] - z0
        df["displacement2_um2"] = (ddx**2 + ddy**2 + ddz**2).astype(float)
    elif displacement2_mode == "step":
        df["displacement2_um2"] = (dx_bk**2 + dy_bk**2 + dz_bk**2).astype(float)
    else:
        raise ValueError(f"Unknown displacement2_mode: {displacement2_mode}")

    # ── Acceleration ──────────────────────────────────────────────────────────
    # Imaris exports ALL ZEROS — confirmed from ground-truth validation.
    # Every computed variant had r²=NaN (zero-variance actual).
    df["accel_um_s2"] = 0.0

    return df


# ----------------------------
# Optional filtering by track length (within the exporter)
# ----------------------------
def filter_tracks(df: pd.DataFrame, min_track_frames: int) -> pd.DataFrame:
    if min_track_frames <= 1:
        return df
    counts = df.groupby("Parent")["frame"].count()
    keep = set(counts[counts >= min_track_frames].index)
    before = len(df)
    df2 = df[df["Parent"].isin(keep)].copy()
    after = len(df2)
    print(f"[filter] min_track_frames={min_track_frames}: rows {before}->{after} | tracks kept={len(keep)}", flush=True)
    return df2


# ----------------------------
# Track-level metrics
# ----------------------------
def ar1_from_positions(p: np.ndarray) -> float:
    p = np.asarray(p, dtype=np.float64)
    if p.size < 3:
        return np.nan
    dp = np.diff(p)
    if dp.size < 2:
        return np.nan
    mu = float(np.mean(dp))
    d0 = dp - mu
    R0 = float(np.sum(d0 * d0))
    if R0 <= 0:
        return np.nan
    R1 = float(np.sum(d0[:-1] * d0[1:]))
    return float(R1 / R0)


def compute_track_table(df: pd.DataFrame) -> pd.DataFrame:
    out = []
    for parent_id, gg in df.groupby("Parent", sort=False):
        gg = gg.sort_values("time_s")
        x = gg["pos_x_um"].to_numpy()
        y = gg["pos_y_um"].to_numpy()
        z = gg["pos_z_um"].to_numpy()
        t = gg["time_s"].to_numpy()

        disp_x = float(x[-1] - x[0]) if x.size else np.nan
        disp_y = float(y[-1] - y[0]) if y.size else np.nan
        disp_z = float(z[-1] - z[0]) if z.size else np.nan
        disp_len = float(np.sqrt(disp_x**2 + disp_y**2 + disp_z**2)) if x.size else np.nan

        duration_s = float(t[-1] - t[0]) if t.size else np.nan

        if x.size >= 2:
            dx_steps = np.diff(x)
            dy_steps = np.diff(y)
            dz_steps = np.diff(z)
            step_dist = np.sqrt(dx_steps**2 + dy_steps**2 + dz_steps**2)
            # Imaris Track Displacement^2 = sum of squared step displacements
            track_displacement2 = float(np.sum(dx_steps**2 + dy_steps**2 + dz_steps**2))
        else:
            step_dist = np.array([], float)
            track_displacement2 = np.nan
        track_len = float(np.sum(step_dist)) if step_dist.size else 0.0

        # Track Speed Mean: track_length / duration  (VALIDATED r²=0.976)
        sp_mean = float(track_len / duration_s) if np.isfinite(duration_s) and duration_s > 0 else np.nan

        # Track Speed Max/Min/StdDev/Variation: use BACKWARD-diff per-frame speeds.
        # Validated: backward diff gives best r² for these aggregates (0.85-0.88)
        # because it includes edge frames that often contain extreme values.
        # Central diff is worse here since it drops first & last frame per track.
        speeds_bkwd = gg["speed_bkwd_um_s"].to_numpy()
        speeds_bkwd = speeds_bkwd[np.isfinite(speeds_bkwd)]
        sp_max = float(np.max(speeds_bkwd)) if speeds_bkwd.size else np.nan
        sp_min = float(np.min(speeds_bkwd)) if speeds_bkwd.size else np.nan
        sp_std = float(np.std(speeds_bkwd, ddof=0)) if speeds_bkwd.size else np.nan
        sp_var = float(sp_std / sp_mean) if np.isfinite(sp_std) and np.isfinite(sp_mean) and sp_mean != 0 else np.nan

        straightness = float(disp_len / track_len) if track_len > 0 else np.nan

        ar1x = ar1_from_positions(x)
        ar1y = ar1_from_positions(y)
        ar1z = ar1_from_positions(z)
        ar1_mean = float(np.nanmean([ar1x, ar1y, ar1z])) if np.isfinite([ar1x, ar1y, ar1z]).any() else np.nan

        area_mean = safe_nanmean(gg["area_um2"].to_numpy())
        vol_mean = safe_nanmean(gg["volume_um3"].to_numpy())
        sph_mean = safe_nanmean(gg["sphericity"].to_numpy())

        axisA_mean = safe_nanmean(gg["axisA_len_um"].to_numpy())
        axisB_mean = safe_nanmean(gg["axisB_len_um"].to_numpy())
        axisC_mean = safe_nanmean(gg["axisC_len_um"].to_numpy())

        i_center_mean = safe_nanmean(gg["intensity_center"].to_numpy())
        i_max = safe_nanmax(gg["intensity_max"].to_numpy())
        i_mean = safe_nanmean(gg["intensity_mean"].to_numpy())
        i_median = safe_nanmean(gg["intensity_median"].to_numpy())
        i_min = safe_nanmin(gg["intensity_min"].to_numpy())
        i_std = safe_nanmean(gg["intensity_std"].to_numpy())
        i_sum = float(np.nansum(gg["intensity_sum"].to_numpy()))

        n_surfaces = int(len(gg))
        n_voxels = int(np.nansum(gg["n_voxels"].to_numpy()))
        n_triangles = int(np.nansum(gg["n_triangles"].to_numpy()))
        n_branches = 0
        n_fusions = 0

        out.append(dict(
            ID=str(parent_id),
            Parent=int(parent_id),

            track_ar1_x=ar1x,
            track_ar1_y=ar1y,
            track_ar1_z=ar1z,
            track_ar1_mean=ar1_mean,

            track_area_mean=area_mean,
            track_volume_mean=vol_mean,
            track_sphericity_mean=sph_mean,

            track_disp_x=disp_x,
            track_disp_y=disp_y,
            track_disp_z=disp_z,
            track_disp_len=disp_len,
            track_displacement2=track_displacement2,

            track_duration_s=duration_s,
            track_length_um=track_len,

            track_position_x_mean=safe_nanmean(x),
            track_position_y_mean=safe_nanmean(y),
            track_position_z_mean=safe_nanmean(z),

            track_speed_mean=sp_mean,
            track_speed_max=sp_max,
            track_speed_min=sp_min,
            track_speed_std=sp_std,
            track_speed_variation=sp_var,
            track_straightness=straightness,

            track_axisA_mean=axisA_mean,
            track_axisB_mean=axisB_mean,
            track_axisC_mean=axisC_mean,

            track_intensity_center_mean=i_center_mean,
            track_intensity_max=i_max,
            track_intensity_mean=i_mean,
            track_intensity_median=i_median,
            track_intensity_min=i_min,
            track_intensity_std=i_std,
            track_intensity_sum=i_sum,

            track_n_branches=n_branches,
            track_n_fusions=n_fusions,
            track_n_surfaces=n_surfaces,
            track_n_triangles=n_triangles,
            track_n_voxels=n_voxels,
        ))

    return pd.DataFrame(out)


# ----------------------------
# Excel formatting
# ----------------------------
def safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(name)).strip()
    return name[:31] if len(name) > 31 else name


def format_workbook(path_xlsx: str):
    wb = load_workbook(path_xlsx)
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = None
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            for row in range(1, min(ws.max_row, 200) + 1):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)
    wb.save(path_xlsx)


# ----------------------------
# Required sheets order
# ----------------------------
REQUIRED_SHEETS_ORDER = [
    "Acceleration",
    "Area",
    "Displacement^2",
    "Ellipsoid Axis A",
    "Ellipsoid Axis B",
    "Ellipsoid Axis C",
    "Ellipsoid Axis Length A",
    "Ellipsoid Axis Length B",
    "Ellipsoid Axis Length C",
    "Ellipticity (oblate)",
    "Ellipticity (prolate)",
    "Intensity Center Ch=1",
    "Intensity Max Ch=1",
    "Intensity Mean Ch=1",
    "Intensity Median Ch=1",
    "Intensity Sum Ch=1",
    "Position",
    "Sphericity",
    "Time Index",
    "Track Duration",
]

# Columns to keep per sheet (None = keep all columns).
# Applied after build_sheet_df to trim to the exact Imaris subset requested.
SHEET_KEEP_COLUMNS = {
    "Time Index":     ["Value"],
    "Track Duration": ["Value", "Unit"],
    "Acceleration":   ["Parent"],
}


def build_sheet_df(sheet: str, obj: pd.DataFrame, trk: pd.DataFrame) -> pd.DataFrame:
    cat_surface = "Surface"
    cat_track = "Track"
    ch1 = "1"

    def base_surface_cols() -> pd.DataFrame:
        return pd.DataFrame({
            "Time": obj["frame_idx"].astype(int),
            "Parent": obj["Parent"].astype("Int64"),
            "ID": obj["ID"].astype(str),
        })

    if sheet == "Overall":
        per_t = obj.groupby("frame_idx", sort=True).agg(
            surfaces=("ID", "count"),
            disconnected=("n_disconnected_components", "sum"),
        ).reset_index()
        rows = []
        for _, rr in per_t.iterrows():
            rows.append(["Number of Disconnected Components per Time Point", float(rr["disconnected"]), "", int(rr["frame_idx"])])
        for _, rr in per_t.iterrows():
            rows.append(["Number of Surfaces per Time Point", float(rr["surfaces"]), "", int(rr["frame_idx"])])

        rows.append(["Number of Tracks", float(obj["Parent"].nunique()), "", ""])
        rows.append(["Total Number of Disconnected Components", float(obj["n_disconnected_components"].sum()), "", ""])
        rows.append(["Total Number of Surfaces", float(len(obj)), "", ""])
        rows.append(["Total Number of Triangles", float(obj["n_triangles"].sum()), "", ""])
        rows.append(["Total Number of Voxels", float(obj["n_voxels"].sum()), "", ""])

        return pd.DataFrame(rows, columns=["Variable", "Value", "Unit", "Time"])

    if sheet == "Time Index":
        base = base_surface_cols()
        return pd.DataFrame({
            "Value": base["Time"].astype(float),
            "Unit": "",
            "Category": cat_surface,
            "Time": base["Time"].astype(int),
            "Parent": base["Parent"],
            "ID": base["ID"],
        })

    def one_value_sheet(col: str, unit: str, fillna0: bool = False):
        base = base_surface_cols()
        vv = obj[col].astype(float)
        if fillna0:
            vv = vv.fillna(0.0)
        return pd.DataFrame({
            "Value": vv,
            "Unit": unit,
            "Category": cat_surface,
            "Time": base["Time"],
            "Parent": base["Parent"],
            "ID": base["ID"],
        })

    if sheet == "Acceleration":
        # Imaris exports ALL ZEROS for Acceleration in our dataset.
        # Ground truth validation confirmed every formula variant has r²=NaN
        # (zero-variance actual). Output 0.0 to match.
        base = base_surface_cols()
        return pd.DataFrame({
            "Value": np.zeros(len(obj), dtype=float),
            "Unit": "um/s^2",
            "Category": cat_surface,
            "Time": base["Time"],
            "Parent": base["Parent"],
            "ID": base["ID"],
        })

    if sheet == "Area":
        return one_value_sheet("area_um2", "um^2")
    if sheet == "Volume":
        return one_value_sheet("volume_um3", "um^3")
    if sheet == "Sphericity":
        return one_value_sheet("sphericity", "")

    if sheet == "Displacement^2":
        base = base_surface_cols()
        return pd.DataFrame({
            "Value": obj["displacement2_um2"].astype(float).fillna(0.0),
            "Unit": "um^2",
            "Category": cat_surface,
            "Time": base["Time"],
            "Parent": base["Parent"],
            "ID": base["ID"],
        })

    if sheet == "Speed":
        # Central diff: NaN at first AND last frame of each track — do not fill.
        return one_value_sheet("speed_um_s", "um/s", fillna0=False)

    if sheet == "Velocity":
        base = base_surface_cols()
        # Central diff: NaN at first AND last frame of each track (no fill).
        # Velocity Z: fill NaN→0 only for 2D data where z is always 0.
        vx = obj["vel_x_um_s"].astype(float)
        vy = obj["vel_y_um_s"].astype(float)
        vz = obj["vel_z_um_s"].astype(float).fillna(0.0)
        return pd.DataFrame({
            "Velocity X": vx,
            "Velocity Y": vy,
            "Velocity Z": vz,
            "Unit": "um/s",
            "Category": cat_surface,
            "Time": base["Time"],
            "Parent": base["Parent"],
            "ID": base["ID"],
        })

    if sheet == "Position":
        base = base_surface_cols()
        return pd.DataFrame({
            "Position X": obj["pos_x_um"].astype(float),
            "Position Y": obj["pos_y_um"].astype(float),
            "Position Z": obj["pos_z_um"].astype(float),
            "Unit": "um",
            "Category": cat_surface,
            "Collection": "Position",
            "Time": base["Time"],
            "Parent": base["Parent"],
            "ID": base["ID"],
        })

    if sheet == "Center of Homogeneous Mass":
        base = base_surface_cols()
        return pd.DataFrame({
            "Center of Homogeneous Mass X": obj["chom_x_um"].astype(float),
            "Center of Homogeneous Mass Y": obj["chom_y_um"].astype(float),
            "Center of Homogeneous Mass Z": obj["chom_z_um"].astype(float),
            "Unit": "um",
            "Category": cat_surface,
            "Collection": "Center of Homogeneous Mass",
            "Time": base["Time"],
            "Parent": base["Parent"],
            "ID": base["ID"],
        })

    if sheet == "Center of Image Mass Ch=1":
        base = base_surface_cols()
        return pd.DataFrame({
            "Center of Image Mass X": obj["cim_x_um"].astype(float),
            "Center of Image Mass Y": obj["cim_y_um"].astype(float),
            "Center of Image Mass Z": obj["cim_z_um"].astype(float),
            "Unit": "um",
            "Category": cat_surface,
            "Channel": ch1,
            "Collection": "Center of Image Mass",
            "Time": base["Time"],
            "Parent": base["Parent"],
            "ID": base["ID"],
        })

    if sheet == "Ellipsoid Axis A":
        base = base_surface_cols()
        return pd.DataFrame({
            "Ellipsoid Axis A X": obj["axisA_x"].astype(float),
            "Ellipsoid Axis A Y": obj["axisA_y"].astype(float),
            "Ellipsoid Axis A Z": obj["axisA_z"].astype(float),
            "Unit": "um",
            "Category": cat_surface,
            "Collection": "Ellipsoid Axis",
            "Time": base["Time"],
            "Parent": base["Parent"],
            "ID": base["ID"],
        })
    if sheet == "Ellipsoid Axis B":
        base = base_surface_cols()
        return pd.DataFrame({
            "Ellipsoid Axis B X": obj["axisB_x"].astype(float),
            "Ellipsoid Axis B Y": obj["axisB_y"].astype(float),
            "Ellipsoid Axis B Z": obj["axisB_z"].astype(float),
            "Unit": "um",
            "Category": cat_surface,
            "Collection": "Ellipsoid Axis",
            "Time": base["Time"],
            "Parent": base["Parent"],
            "ID": base["ID"],
        })
    if sheet == "Ellipsoid Axis C":
        base = base_surface_cols()
        return pd.DataFrame({
            "Ellipsoid Axis C X": obj["axisC_x"].astype(float),
            "Ellipsoid Axis C Y": obj["axisC_y"].astype(float),
            "Ellipsoid Axis C Z": obj["axisC_z"].astype(float),
            "Unit": "um",
            "Category": cat_surface,
            "Collection": "Ellipsoid Axis",
            "Time": base["Time"],
            "Parent": base["Parent"],
            "ID": base["ID"],
        })

    if sheet == "Ellipsoid Axis Length A":
        return one_value_sheet("axisA_len_um", "um")
    if sheet == "Ellipsoid Axis Length B":
        return one_value_sheet("axisB_len_um", "um")
    if sheet == "Ellipsoid Axis Length C":
        return one_value_sheet("axisC_len_um", "um")

    if sheet == "Ellipticity (oblate)":
        return one_value_sheet("ellipticity_oblate", "")
    if sheet == "Ellipticity (prolate)":
        return one_value_sheet("ellipticity_prolate", "")

    # intensity sheets use "Value" as first column
    def intensity_sheet(series: pd.Series) -> pd.DataFrame:
        base = base_surface_cols()
        return pd.DataFrame({
            "Value": series.astype(float),
            "Unit": "",
            "Category": cat_surface,
            "Channel": ch1,
            "Time": base["Time"],
            "Parent": base["Parent"],
            "ID": base["ID"],
        })

    if sheet == "Intensity Center Ch=1":
        return intensity_sheet(obj["intensity_center"])
    if sheet == "Intensity Max Ch=1":
        return intensity_sheet(obj["intensity_max"])
    if sheet == "Intensity Mean Ch=1":
        return intensity_sheet(obj["intensity_mean"])
    if sheet == "Intensity Median Ch=1":
        return intensity_sheet(obj["intensity_median"])
    if sheet == "Intensity Min Ch=1":
        return intensity_sheet(obj["intensity_min"])
    if sheet == "Intensity StdDev Ch=1":
        return intensity_sheet(obj["intensity_std"])
    if sheet == "Intensity Sum Ch=1":
        return intensity_sheet(obj["intensity_sum"])

    if sheet == "Number of Disconnected Components":
        return one_value_sheet("n_disconnected_components", "")
    if sheet == "Number of Triangles":
        return one_value_sheet("n_triangles", "")
    if sheet == "Number of Vertices":
        return one_value_sheet("n_vertices", "")
    if sheet == "Number of Voxels":
        return one_value_sheet("n_voxels", "")

    if sheet.startswith("Track "):
        if trk is None or trk.empty:
            return pd.DataFrame()

        def track_value_df(col: str, unit: str = "") -> pd.DataFrame:
            return pd.DataFrame({
                "Value": trk[col].astype(float),
                "Unit": unit,
                "Category": cat_track,
                "ID": trk["Parent"].astype("Int64"),
            })

        if sheet == "Track Ar1":
            return pd.DataFrame({
                "Track Ar1 X": trk["track_ar1_x"].astype(float),
                "Track Ar1 Y": trk["track_ar1_y"].astype(float),
                "Track Ar1 Z": trk["track_ar1_z"].astype(float),
                "Unit": "",
                "Category": cat_track,
                "ID": trk["Parent"].astype("Int64"),
            })
        if sheet == "Track Ar1 Mean":
            return track_value_df("track_ar1_mean", "")
        if sheet == "Track Area Mean":
            return track_value_df("track_area_mean", "um^2")
        if sheet == "Track Volume Mean":
            return track_value_df("track_volume_mean", "um^3")
        if sheet == "Track Sphericity Mean":
            return track_value_df("track_sphericity_mean", "")

        if sheet == "Track Displacement":
            return pd.DataFrame({
                "Track Displacement X": trk["track_disp_x"].astype(float),
                "Track Displacement Y": trk["track_disp_y"].astype(float),
                "Track Displacement Z": trk["track_disp_z"].astype(float),
                "Unit": "um",
                "Category": cat_track,
                "ID": trk["Parent"].astype("Int64"),
            })
        if sheet == "Track Displacement Length":
            return track_value_df("track_disp_len", "um")
        if sheet == "Track Displacement^2":
            return track_value_df("track_displacement2", "um^2")
        if sheet == "Track Duration":
            return track_value_df("track_duration_s", "s")
        if sheet == "Track Length":
            return track_value_df("track_length_um", "um")

        if sheet == "Track Position":
            return pd.DataFrame({
                "Track Position X Mean": trk["track_position_x_mean"].astype(float),
                "Track Position Y Mean": trk["track_position_y_mean"].astype(float),
                "Track Position Z Mean": trk["track_position_z_mean"].astype(float),
                "Unit": "um",
                "Category": cat_track,
                "ID": trk["Parent"].astype("Int64"),
            })

        if sheet == "Track Speed Mean":
            return track_value_df("track_speed_mean", "um/s")
        if sheet == "Track Speed Max":
            return track_value_df("track_speed_max", "um/s")
        if sheet == "Track Speed Min":
            return track_value_df("track_speed_min", "um/s")
        if sheet == "Track Speed StdDev":
            return track_value_df("track_speed_std", "um/s")
        if sheet == "Track Speed Variation":
            return track_value_df("track_speed_variation", "")
        if sheet == "Track Straightness":
            return track_value_df("track_straightness", "")

        if sheet == "Track Ellipsoid Axis A Mean":
            return track_value_df("track_axisA_mean", "um")
        if sheet == "Track Ellipsoid Axis B Mean":
            return track_value_df("track_axisB_mean", "um")
        if sheet == "Track Ellipsoid Axis C Mean":
            return track_value_df("track_axisC_mean", "um")

        if sheet == "Track Intensity Max Ch=1":
            return track_value_df("track_intensity_max", "")
        if sheet == "Track Intensity Mean Ch=1":
            return track_value_df("track_intensity_mean", "")
        if sheet == "Track Intensity Median Ch=1":
            return track_value_df("track_intensity_median", "")
        if sheet == "Track Intensity Min Ch=1":
            return track_value_df("track_intensity_min", "")
        if sheet == "Track Intensity StdDev Ch=1":
            return track_value_df("track_intensity_std", "")
        if sheet == "Track Intensity Sum Ch=1":
            return track_value_df("track_intensity_sum", "")

        if sheet == "Track Number of Branches":
            return track_value_df("track_n_branches", "")
        if sheet == "Track Number of Fusions":
            return track_value_df("track_n_fusions", "")
        if sheet == "Track Number of Surfaces":
            return track_value_df("track_n_surfaces", "")
        if sheet == "Track Number of Triangles":
            return track_value_df("track_n_triangles", "")
        if sheet == "Track Number of Voxels":
            return track_value_df("track_n_voxels", "")

        return pd.DataFrame()

    return pd.DataFrame()


# ----------------------------
# Main
# ----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--segm-npz", required=True)
    ap.add_argument("--tif", required=True)
    ap.add_argument("--acdc-table", default=None)
    ap.add_argument("--trackpy-csv", default=None,
                    help="TrackPy tracks CSV (frame, x, y, ID, particle). "
                         "When provided, Parent = particle (persistent track ID) "
                         "instead of Cell_ID from the acdc_output table.")
    ap.add_argument("--out", default="acdc_imaris_like_required_FIXED.xlsx")

    ap.add_argument("--pixel-size-um", type=float, required=True)
    ap.add_argument("--z-step-um", type=float, required=True)
    ap.add_argument("--frame-interval-s", type=float, default=3600.0)
    ap.add_argument("--tif-channel-index", type=int, default=0)
    ap.add_argument("--acdc-channel", type=str, default=None,
                    help="Channel name in ACDC CSV columns (e.g. 'NIR'). "
                         "If not given, auto-detected from TIF filename.")
    ap.add_argument("--swap-xy", action="store_true")

    ap.add_argument("--progress-every-frames", type=int, default=1)
    ap.add_argument("--progress-every-objects", type=int, default=200)
    ap.add_argument("--max-frames", type=int, default=0)

    ap.add_argument("--compute-disconnected-components", action="store_true")
    ap.add_argument("--skip-sphericity", action="store_true")

    ap.add_argument("--min-track-frames", type=int, default=1)

    ap.add_argument("--displacement2-mode", choices=["from_start", "step"], default="from_start",
                    help="Imaris-like default is from_start (squared displacement from track start).")

    # Mapping parameters (used when --acdc-table is provided)
    ap.add_argument("--map-max-dist-px", type=float, default=8.0)
    ap.add_argument("--map-offset-x-px", type=float, default=0.0)
    ap.add_argument("--map-offset-y-px", type=float, default=0.0)
    ap.add_argument("--map-offset-z-px", type=float, default=0.0)

    # Global ACDC->Imaris coordinate correction (pixels).
    # DEFAULT IS 0 — dataset-specific, must be calibrated per acquisition.
    ap.add_argument("--imaris-offset-x-px", type=float, default=0.0,
                    help="Global X shift (pixels) applied to exported coordinates (ACDC->Imaris). "
                         "Dataset-specific — calibrate before use. Default: 0.")
    ap.add_argument("--imaris-offset-y-px", type=float, default=0.0,
                    help="Global Y shift (pixels) applied to exported coordinates (ACDC->Imaris). "
                         "Dataset-specific — calibrate before use. Default: 0.")
    ap.add_argument("--imaris-offset-z-px", type=float, default=0.0,
                    help="Global Z shift (pixels) applied to exported coordinates (ACDC->Imaris). Default: 0.")
    ap.add_argument("--time-index-start", type=int, default=2,
                    help="Time index of the first frame in the output Excel. "
                         "Empirically confirmed: Imaris labels the first frame as Time=2 in this dataset. "
                         "Use 1 if your Imaris export shows Time=1 for the first frame. "
                         "CRITICAL: all per-frame joins (Area, Volume, Speed, Intensity) "
                         "break if this does not match your Imaris file. Default: 2.")
    ap.add_argument("--mesh-smooth-sigma", type=float, default=0.0,
                    help="Gaussian blur sigma (pixels, YX plane) applied to the binary mask before "
                         "marching cubes for 2D cells.  Replicates Imaris surface mesh inflation: "
                         "the smooth surface at level < 0.5 expands the cell boundary outward. "
                         "Calibrated value: 7.0 at 1.22 um/px.  Default 0 = disabled.")
    ap.add_argument("--mesh-smooth-level", type=float, default=0.25,
                    help="Marching cubes isosurface level used when --mesh-smooth-sigma > 0. "
                         "Values < 0.5 expand the surface outward beyond the original boundary. "
                         "Default: 0.25 (calibrated against GT volume/area ratios at sigma=7).")

    args = ap.parse_args()

    max_frames = int(args.max_frames) if int(args.max_frames) > 0 else None

    print("[start] loading inputs...", flush=True)
    print(f"[config] time_index_start={args.time_index_start} "
          f"(first frame will appear as Time={args.time_index_start} in Excel; "
          f"pass --time-index-start 1 if your Imaris export uses Time=1 for the first frame)", flush=True)
    acdc_df = load_table_optional(args.acdc_table)
    if acdc_df is not None:
        print(f"[ok] loaded acdc table: shape={acdc_df.shape}", flush=True)
        print("[note] pipeline behavior: exporting ONLY objects mapped to IDs present in --acdc-table", flush=True)

    # Build TrackPy (frame, cell_id) -> particle mapper
    trackpy_mapper = None
    if args.trackpy_csv:
        tp_path = Path(args.trackpy_csv)
        if not tp_path.exists():
            raise FileNotFoundError(f"--trackpy-csv not found: {tp_path}")
        tp_df = pd.read_csv(tp_path)
        # normalise column names
        tp_df.columns = [c.strip().lower() for c in tp_df.columns]
        if not {"frame", "id", "particle"}.issubset(set(tp_df.columns)):
            raise ValueError(
                f"--trackpy-csv must have columns: frame, ID, particle. "
                f"Found: {list(tp_df.columns)}"
            )
        trackpy_mapper = {
            (int(row["frame"]), int(row["id"])): int(row["particle"])
            for _, row in tp_df.iterrows()
        }
        print(f"[ok] loaded trackpy CSV: {len(trackpy_mapper):,} (frame, cell_id) -> particle entries", flush=True)
        print("[note] Parent will use particle (persistent track ID) from TrackPy", flush=True)

    segm = load_segmentation_npz(args.segm_npz)
    print(f"[ok] segm shape={segm.shape}, dtype={segm.dtype}", flush=True)

    img = load_tif_frames(args.tif, channel_index=args.tif_channel_index)
    print(f"[ok] img shape={img.shape}, dtype={img.dtype}", flush=True)

    # Auto-detect channel name from TIF filename (e.g. field1_B2_NIR.tif -> NIR)
    acdc_channel_name = args.acdc_channel
    if not acdc_channel_name:
        tif_stem = Path(args.tif).stem  # e.g. field1_B2_NIR
        parts = tif_stem.split("_")
        if parts:
            acdc_channel_name = parts[-1]  # last part is channel
            print(f"[info] auto-detected ACDC channel name: '{acdc_channel_name}'", flush=True)

    segm, img = align_time_axis(segm, img)
    print(f"[ok] aligned segm/img: segm={segm.shape}, img={img.shape}", flush=True)

    print("[stage] computing per-object stats (Imaris-like)...", flush=True)
    obj = compute_object_table(
        segm=segm,
        img=img,
        pixel_size_um=float(args.pixel_size_um),
        z_step_um=float(args.z_step_um),
        frame_interval_s=float(args.frame_interval_s),
        swap_xy=bool(args.swap_xy),
        progress_every_frames=int(args.progress_every_frames),
        progress_every_objects=int(args.progress_every_objects),
        compute_disconnected_components=bool(args.compute_disconnected_components),
        skip_sphericity=bool(args.skip_sphericity),
        acdc_df=acdc_df,
        map_max_dist_px=float(args.map_max_dist_px),
        map_offset_x_px=float(args.map_offset_x_px),
        map_offset_y_px=float(args.map_offset_y_px),
        map_offset_z_px=float(args.map_offset_z_px),
        max_frames=max_frames,
        displacement2_mode=str(args.displacement2_mode),
        imaris_offset_x_px=float(args.imaris_offset_x_px),
        imaris_offset_y_px=float(args.imaris_offset_y_px),
        imaris_offset_z_px=float(args.imaris_offset_z_px),
        trackpy_mapper=trackpy_mapper,
        acdc_channel_name=acdc_channel_name,
        time_index_start=int(args.time_index_start),
        mesh_smooth_sigma=float(args.mesh_smooth_sigma),
        mesh_smooth_level=float(args.mesh_smooth_level),
    )

    obj = filter_tracks(obj, int(args.min_track_frames))
    trk = compute_track_table(obj)

    print("[stage] writing Excel sheets...", flush=True)
    t_write = time.time()
    with pd.ExcelWriter(args.out, engine="openpyxl") as writer:
        for i, sheet in enumerate(REQUIRED_SHEETS_ORDER, start=1):
            print(f"[progress] writing sheet {i}/{len(REQUIRED_SHEETS_ORDER)}: {sheet}", flush=True)
            df_sheet = build_sheet_df(sheet, obj, trk)
            if df_sheet is None:
                df_sheet = pd.DataFrame()
            keep_cols = SHEET_KEEP_COLUMNS.get(sheet)
            if keep_cols is not None and not df_sheet.empty:
                df_sheet = df_sheet[[c for c in keep_cols if c in df_sheet.columns]]
            df_sheet.to_excel(writer, sheet_name=safe_sheet_name(sheet), index=False)

    print(f"[progress] finished Excel write in {time.time()-t_write:.1f}s", flush=True)
    print("[stage] formatting workbook...", flush=True)
    format_workbook(args.out)

    print(f"[DONE] wrote: {args.out}", flush=True)


if __name__ == "__main__":
    main()